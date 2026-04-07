from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# SummitSpan Builders Group — Mission, Vision, Values

Mission: Build projects safely with predictable delivery and transparent job-costing.
Vision: The regional GC owners trust for execution at scale.
Values: Safety, schedule discipline, documentation.
""",
    )

    w(
        "brand-guide.md",
        """# SummitSpan — Brand Guide (v0.1)

Palette:
- Navy: #0B1F3A
- Orange: #F97316
- Off-white: #F7F5F0

Voice: professional, direct.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Marcus Lee

- Reduce schedule drift (single source of truth)
- Tighten CO discipline (approved vs pending)
- Align PM forecast with WIP
""",
    )

    w(
        "project-plan-riverwalk.md",
        """# Project Plan — Multifamily Riverwalk (RW-4412)

- Scope: multifamily build
- Risk: long-lead windows
- Baseline: substantial completion mid-July (may slip)
""",
    )

    w(
        "meeting-notes-riverwalk-2026-05-08.md",
        """# Weekly Meeting Notes — Riverwalk — 2026-05-08

- Schedule on track
- COs: 2 pending
""",
    )

    w(
        "meeting-notes-riverwalk-2026-06-05.md",
        """# Weekly Meeting Notes — Riverwalk — 2026-06-05

- Substantial completion target (per notes): 2026-07-15
- Long-lead windows slipping
""",
    )

    w(
        "meeting-notes-riverwalk-2026-07-03.md",
        """# Weekly Meeting Notes — Riverwalk — 2026-07-03

- Re-baseline being finalized
- CO totals need clarity
""",
    )

    w(
        "exec-summary-email-2026-07-04.md",
        """# Exec Summary Email — 2026-07-04

- Substantial completion: early Aug (per master schedule)
- COs: $2.6M including pending
- Margin: WIP suggests compression
""",
    )

    w(
        "safety-walk-template.md",
        """# Safety Walk Template

Checks: PPE, housekeeping, fall protection, hot work.
""",
    )

    w(
        "incident-report-template.md",
        """# Incident Report Template

Date:
Project:
Incident:
Actions:
Owner:
""",
    )

    w(
        "subcontractor-evaluation-template.md",
        """# Subcontractor Evaluation Template

Quality, schedule, safety, communication.
""",
    )

    w(
        "rfi-template.md",
        """# RFI Template

RFI #:
Question:
Needed by:
Owner:
""",
    )


def make_json():
    wjson(
        "org-structure.json",
        {
            "generated": str(date.today()),
            "leaders": {
                "CEO": "Dana Whitfield",
                "COO": "Marcus Lee",
                "Precon": "Priya Nair",
                "Ops": "Elena Ortiz",
                "Controller": "Sam Cho",
                "Safety": "Tasha Reed",
            },
            "offices": 4,
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-06-01",
            "kpis": [
                {"name": "CO total (approved)", "definition": "sum approved COs"},
                {"name": "CO total (forecast)", "definition": "approved + pending"},
                {"name": "Margin (PM)", "definition": "forecast fee %"},
                {"name": "Margin (WIP)", "definition": "WIP margin %"},
            ],
        },
    )

    wjson(
        "owner-change-request-form.json",
        {
            "schema_version": "0.1",
            "title": "Owner Change Request",
            "fields": [
                {"id": "project", "type": "string", "required": True},
                {"id": "description", "type": "string", "required": True},
                {"id": "cost_impact", "type": "number"},
                {"id": "schedule_impact_days", "type": "integer"},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "company-capabilities-one-pager.pdf",
        "SummitSpan Builders — Capabilities",
        ["Commercial", "Multifamily", "Light civil", "Service division"],
    )

    pdf(
        BASE / "sample-estimate-multifamily-riverwalk.pdf",
        "Sample Estimate — Riverwalk",
        ["Base GMP: $48.2M", "Contingency: $1.6M", "Long-lead windows excluded"],
    )

    pdf(
        BASE / "contract-summary-multifamily-riverwalk.pdf",
        "Contract Summary — Riverwalk",
        ["Contract type: GMP", "COs tracked approved vs pending", "Baseline schedule + re-baselines"],
    )


def make_xlsx():
    import openpyxl

    # estimate template
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Estimate"
    ws.append(["Division","Description","Amount"])
    for r in [("01","General conditions",6800000),("03","Concrete",9200000),("06","Framing",8700000),("09","Finishes",11200000)]:
        ws.append(list(r))
    for i in range(2, ws.max_row+1): ws.cell(i,3).number_format="$#,##0"
    wb.save(BASE/"estimate-template.xlsx")

    # master schedule
    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Master"
    ws2.append(["Milestone","Date","Notes"])
    ws2.append(["Substantial Completion","2026-08-02","re-baselined"])
    wb2.save(BASE/"master-schedule-riverwalk.xlsx")

    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="Lookahead"
    ws3.append(["Week","Focus","Owner"])
    ws3.append(["2026-W27","Window install prep","Super"])
    wb3.save(BASE/"lookahead-schedule-riverwalk.xlsx")

    # RFI log
    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="RFIs"
    ws4.append(["RFI #","Date","Question","Status","Owner"])
    ws4.append(["RFI-112","2026-06-01","Window spec clarification","Open","PM"])
    wb4.save(BASE/"rfi-log-riverwalk.xlsx")

    # Submittal log
    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="Submittals"
    ws5.append(["Submittal","Date","Status","Owner"])
    ws5.append(["Window shop drawings","2026-05-20","Revise","Super"])
    wb5.save(BASE/"submittal-log-riverwalk.xlsx")

    # CO log
    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="COs"
    ws6.append(["CO #","Date","Description","Cost","Status"])
    for r in [("CO-14","2026-05-30","Design change",650000,"Approved"),("CO-18","2026-06-22","Window upgrade",1450000,"Approved"),("CO-21","2026-06-30","Sitework add",500000,"Pending")]:
        ws6.append(list(r))
    for i in range(2, ws6.max_row+1): ws6.cell(i,4).number_format="$#,##0"
    wb6.save(BASE/"change-order-log-riverwalk.xlsx")

    # Job costing and WIP
    wb7=openpyxl.Workbook(); ws7=wb7.active; ws7.title="Job Cost"
    ws7.append(["Category","Budget","Cost to date","Forecast","Variance"])
    for r in [("Concrete",9200000,6100000,9800000,None),("Framing",8700000,5400000,9100000,None)]:
        ws7.append(list(r))
    for i in range(2, ws7.max_row+1):
        ws7.cell(i,5).value=f"=B{i}-D{i}"
        for c in [2,3,4,5]: ws7.cell(i,c).number_format="$#,##0"
    wb7.save(BASE/"job-costing-riverwalk.xlsx")

    wb8=openpyxl.Workbook(); ws8=wb8.active; ws8.title="WIP"
    ws8.append(["Project","Contract","CO Approved","Cost to Date","Est Cost at Comp","Margin % (WIP)"])
    ws8.append(["Riverwalk",48200000,2100000,31800000,46200000,None])
    ws8.cell(2,6).value="=(B2+C2-E2)/(B2+C2)"; ws8.cell(2,6).number_format="0.0%"
    for c in [2,3,4,5]: ws8.cell(2,c).number_format="$#,##0"
    wb8.save(BASE/"wip-report-2026-06.xlsx")

    wb9=openpyxl.Workbook(); ws9=wb9.active; ws9.title="AR"
    ws9.append(["Client","Invoice","Invoice Date","Amount","Status"])
    ws9.append(["Riverwalk Owner","APP-06","2026-06-30",3200000,"Open"])
    ws9.cell(2,4).number_format="$#,##0"
    wb9.save(BASE/"accounts-receivable-2026-06.xlsx")

    # revenue
    wb10=openpyxl.Workbook(); ws10=wb10.active; ws10.title="Quarterly"
    ws10.append(["Year","Quarter","Revenue"])
    for r in [(2024,"Q4",98000000),(2025,"Q1",94000000),(2025,"Q2",101000000),(2025,"Q3",99000000),(2025,"Q4",112000000)]:
        ws10.append(list(r))
    for i in range(2, ws10.max_row+1): ws10.cell(i,3).number_format="$#,##0"
    wb10.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb11=openpyxl.Workbook(); ws11=wb11.active; ws11.title="2026"
    ws11.append(["Month","Revenue"])
    for r in [("2026-01",36000000),("2026-02",34000000),("2026-03",38000000)]:
        ws11.append(list(r))
    for i in range(2, ws11.max_row+1): ws11.cell(i,2).number_format="$#,##0"
    wb11.save(BASE/"projected-revenue-2026.xlsx")

    # Safety incidents + punchlist
    wb12=openpyxl.Workbook(); ws12=wb12.active; ws12.title="2026"
    ws12.append(["Date","Project","Incident","Severity"])
    ws12.append(["2026-05-12","Riverwalk","Near miss","Low"])
    wb12.save(BASE/"safety-incident-log-2026.xlsx")

    wb13=openpyxl.Workbook(); ws13=wb13.active; ws13.title="Template"
    ws13.append(["Area","Item","Owner","Status"])
    ws13.append(["Unit 304","Paint touch-up","Sub","Open"])
    wb13.save(BASE/"qaqc-punchlist-template.xlsx")


def make_pptx():
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"program-roadmap-2026.pptx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx(); make_pptx()


if __name__ == "__main__":
    main()
