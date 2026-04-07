from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Titan River Components — Mission, Vision, Values

Mission: Build components customers can trust under tight tolerances.
Vision: A multi-plant operation with predictable quality and delivery.
Values: Safety, definitions, root cause discipline.
""",
    )

    w(
        "brand-guide.md",
        """# Titan River Components — Brand Guide (v0.1)

Palette:
- Steel gray: #2F3A44
- Blue: #2563EB
- Safety orange: #F97316
- White: #F8FAFC

Voice: technical, direct, calm.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Erica Shaw

Priorities:
- Align KPI definitions (OEE, scrap, PPM)
- Reduce CAPA backlog
- Tighten vendor risk controls
""",
    )

    w(
        "control-plan-widget-line-v3.md",
        """# Control Plan — Widget Line (v3)

Effective: 2025-02-01

- Sample frequency: every 2 hours
- Rework tracked separately
""",
    )

    w(
        "control-plan-widget-line-v4.md",
        """# Control Plan — Widget Line (v4 draft)

Drafted: 2025-10-18

Changes:
- Sample frequency: every 1 hour
- Rework included in scrap reporting

Known issue: Plants not aligned on rollout.
""",
    )

    w(
        "meeting-notes-program-review-2026-01-20.md",
        """# Program Review — 2026-01-20

- CAPA backlog flagged.
- Plant B downtime higher than expected.
""",
    )

    w(
        "meeting-notes-ops-review-2026-02-06.md",
        """# Ops Review — 2026-02-06

- OEE disagreement persists (plant vs finance).
- Expedite tracker maintained outside ERP.
""",
    )

    w(
        "nonconformance-report-template.md",
        """# Nonconformance Report Template

Date:
Part:
Issue:
Containment:
Root cause:
Corrective action:
Owner:
""",
    )

    w(
        "supplier-evaluation-template.md",
        """# Supplier Evaluation Template

Supplier:
Scores: lead time, quality, pricing.
Notes:
""",
    )

    w(
        "change-request-template.md",
        """# Change Request Template

Change ID:
Description:
Reason:
Risk:
Approvals:
""",
    )


def make_json():
    wjson(
        "org-structure.json",
        {
            "generated": str(date.today()),
            "leaders": {
                "COO": "Erica Shaw",
                "SupplyChain": "Victor Chen",
                "Quality": "Amina Patel",
                "Finance": "Jonah Brooks",
            },
            "plants": ["Plant A", "Plant B", "Plant C"],
        },
    )

    wjson(
        "vendor-list.json",
        {
            "generated": str(date.today()),
            "vendors": [
                {"id": "V001", "name": "GreatLakes Steel", "category": "steel", "risk": "lead time"},
                {"id": "V002", "name": "Midwest Plating", "category": "plating", "risk": "quality"},
                {"id": "V003", "name": "RiverPack", "category": "packaging", "risk": "low"},
            ],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-02-01",
            "kpis": [
                {"name": "OEE (plant)", "definition": "availability excludes planned maintenance"},
                {"name": "OEE (finance)", "definition": "availability includes planned maintenance"},
                {"name": "Scrap (ops)", "definition": "scrap incl rework"},
                {"name": "Scrap (QA)", "definition": "true scrap only"},
                {"name": "OTD (customer)", "definition": "on-time if first shipment arrives"},
                {"name": "OTD (internal)", "definition": "on-time if full order complete"},
            ],
        },
    )

    wjson(
        "expedite-request-form.json",
        {
            "schema_version": "0.1",
            "title": "Expedite Request",
            "fields": [
                {"id": "customer", "type": "string", "required": True},
                {"id": "part", "type": "string", "required": True},
                {"id": "qty", "type": "integer", "required": True},
                {"id": "need_by", "type": "date", "required": True},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "customer-scorecard-2025-q4.pdf",
        "Customer Scorecard — Q4 2025 (Summary)",
        [
            "OTD (customer definition): 96%",
            "Quality PPM: 420 (definition varies)",
            "Partial shipments counted as on-time if first shipment meets requested date.",
        ],
    )

    pdf(
        BASE / "board-update-2026-02.pdf",
        "Board Update — Feb 2026 (Summary)",
        [
            "OEE: 72% (finance definition)",
            "CAPA backlog: elevated",
            "Vendor risk: resin/steel lead times volatile",
        ],
    )


def make_xlsx():
    import openpyxl

    # Production plan
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="H1 2026"
    ws.append(["Week","Plant","Line","Primary Part","Planned Qty","Notes"])
    for r in [("2026-W02","Plant A","Line 2","WIDGET-01",120000,""),("2026-W03","Plant B","Line 1","WIDGET-02",98000,"downtime")]:
        ws.append(list(r))
    wb.save(BASE/"production-plan-2026-h1.xlsx")

    # Downtime log
    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Plant A 2026-01"
    ws2.append(["Date","Line","Minutes","Reason","Notes"])
    for r in [("2026-01-12","Line 2",180,"jam",""),("2026-01-19","Line 2",240,"maintenance","planned?")]:
        ws2.append(list(r))
    wb2.save(BASE/"downtime-log-plant-a-2026-01.xlsx")

    # OEE dashboard contradiction
    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="2026-01"
    ws3.append(["Plant","OEE (plant)","OEE (finance)","Notes"])
    for r in [("Plant A",0.79,0.73,"maintenance assumption"),("Plant B",0.76,0.70,"downtime")]:
        ws3.append(list(r))
    for i in range(2, ws3.max_row+1):
        ws3.cell(i,2).number_format="0.0%"; ws3.cell(i,3).number_format="0.0%"
    wb3.save(BASE/"oee-dashboard-2026-01.xlsx")

    # Plant scorecard
    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="2026-01"
    ws4.append(["Plant","OTD (customer)","OTD (internal)","Scrap (ops)","Scrap (QA)"])
    for r in [("Plant A",0.96,0.92,0.041,0.031),("Plant B",0.95,0.90,0.048,0.036)]:
        ws4.append(list(r))
    for i in range(2, ws4.max_row+1):
        for c in [2,3,4,5]: ws4.cell(i,c).number_format="0.0%"
    wb4.save(BASE/"plant-scorecard-2026-01.xlsx")

    # Safety incidents
    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="2025"
    ws5.append(["Date","Plant","Incident","Severity","Notes"])
    for r in [("2025-09-14","Plant B","Slip incident","Low",""),("2025-12-02","Plant C","Forklift near-miss","Med","")]:
        ws5.append(list(r))
    wb5.save(BASE/"safety-incident-log-2025.xlsx")

    # QA incidents
    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="2025"
    ws6.append(["Date","Plant","Part","Issue","Severity","Notes"])
    for r in [("2025-11-18","Plant A","WIDGET-01","PPM spike","High",""),("2025-12-09","Plant B","WIDGET-02","Rework backlog","Med","")]:
        ws6.append(list(r))
    wb6.save(BASE/"qa-incidents-log-2025.xlsx")

    # CAPA tracker
    wb7=openpyxl.Workbook(); ws7=wb7.active; ws7.title="2026"
    ws7.append(["CAPA","Opened","Owner","Due","Status","Notes"])
    for r in [("CAPA-102","2026-01-05","Amina","2026-02-15","Open","overdue risk"),("CAPA-099","2025-12-12","Luis","2026-01-31","Overdue","" )]:
        ws7.append(list(r))
    wb7.save(BASE/"capa-tracker-2026.xlsx")

    # Vendor risk register
    wb8=openpyxl.Workbook(); ws8=wb8.active; ws8.title="Vendors"
    ws8.append(["Vendor","Category","Risk","Notes"])
    for r in [("GreatLakes Steel","steel","Med","lead time"),("Midwest Plating","plating","Med","quality variability")]:
        ws8.append(list(r))
    wb8.save(BASE/"vendor-risk-register.xlsx")

    # Expedite tracker
    wb9=openpyxl.Workbook(); ws9=wb9.active; ws9.title="2026-02"
    ws9.append(["Date","Customer","Part","Qty","Need by","Status","Notes"])
    for r in [("2026-02-03","AutoCo","WIDGET-01",20000,"2026-02-20","Open","tracked outside ERP")]:
        ws9.append(list(r))
    wb9.save(BASE/"expedite-tracker-2026-02.xlsx")

    # Lead time dashboard
    wb10=openpyxl.Workbook(); ws10=wb10.active; ws10.title="2026-01"
    ws10.append(["Vendor","Promised LT days","Actual LT days","Notes"])
    ws10.append(["GreatLakes Steel",14,18,"winter delays"])
    wb10.save(BASE/"lead-time-dashboard-2026-01.xlsx")

    # Revenue history
    wb11=openpyxl.Workbook(); ws11=wb11.active; ws11.title="Quarterly"
    ws11.append(["Year","Quarter","Revenue","Notes"])
    for r in [(2024,"Q4",32000000,""),(2025,"Q1",30500000,""),(2025,"Q2",34000000,""),(2025,"Q3",35500000,""),(2025,"Q4",37000000,"")]:
        ws11.append(list(r))
    for i in range(2, ws11.max_row+1): ws11.cell(i,3).number_format="$#,##0"
    wb11.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb12=openpyxl.Workbook(); ws12=wb12.active; ws12.title="2026"
    ws12.append(["Month","Revenue","Notes"])
    for r in [("2026-01",11800000,""),("2026-02",11200000,""),("2026-03",12100000,"")]:
        ws12.append(list(r))
    for i in range(2, ws12.max_row+1): ws12.cell(i,2).number_format="$#,##0"
    wb12.save(BASE/"projected-revenue-2026.xlsx")

    # Inventory valuation
    wb13=openpyxl.Workbook(); ws13=wb13.active; ws13.title="2025-12"
    ws13.append(["Item","Qty","Std Cost","Last Purchase","Valuation (ERP)","Valuation (Plant)"])
    ws13.append(["Steel",420000,0.88,0.94,None,None])
    ws13.cell(2,5).value="=B2*C2"; ws13.cell(2,6).value="=B2*D2"
    wb13.save(BASE/"inventory-valuation-2025-12.xlsx")


def make_pptx():
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"program-roadmap-2026.pptx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx(); make_pptx()


if __name__ == "__main__":
    main()
