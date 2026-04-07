from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Northpoint Packaging & Plastics — Mission, Vision, Values

Mission: Reliable packaging components with predictable lead times.
Vision: Be the default supplier for repeat mid-market customers.
Values: Safety, quality, document the work.
""",
    )

    w(
        "brand-guide.md",
        """# Northpoint Packaging & Plastics — Brand Guide (v0.1)

Palette:
- Steel gray: #2F3A44
- Blue: #2563EB
- White: #F8FAFC

Voice: practical, technical when needed.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Simone Grant

- Reduce scrap + rework
- Align definitions across QA/Ops
- Reconcile inventory valuation methods
""",
    )

    w(
        "sop-line-changeover-v2.md",
        """# SOP — Line Changeover (v2)

Effective: 2025-06-01

- Setup checklist (paper)
- First-article sample n=5
""",
    )

    w(
        "sop-line-changeover-v3.md",
        """# SOP — Line Changeover (v3 draft)

Drafted: 2026-01-20

Changes:
- First-article sample n=8
- Record in shared sheet

Known issue: floor still uses v2 printouts.
""",
    )

    w(
        "weekly-ops-review-2026-01-06.md",
        """# Weekly Ops Review — 2026-01-06

- Scrap rate (ops): 4.2%
- QA scrap (true scrap): 3.1% (excludes rework)
""",
    )

    w(
        "weekly-ops-review-2026-02-03.md",
        """# Weekly Ops Review — 2026-02-03

- Inventory adjustments tracked in side sheet.
- Partial shipments affecting OTD.
""",
    )

    w(
        "voice-note-transcript-2026-02-04.md",
        """# Voice Note Transcript — 2026-02-04

- QA and ops scrap numbers differ because of rework.
- Finance uses standard cost; ops keeps using last resin purchase price.
""",
    )

    w(
        "supplier-evaluation-template.md",
        """# Supplier Evaluation Template

Supplier:
Scores: lead time, quality, pricing.
Notes:
""",
    )

    w(
        "nonconformance-report-template.md",
        """# Nonconformance Report Template

Date:
Part:
Issue:
Root cause:
Fix:
""",
    )


def make_json():
    wjson(
        "top-customers.json",
        {
            "generated": str(date.today()),
            "customers": [
                {"account": "Lakefront Foods", "segment": "food"},
                {"account": "CedarMed Devices", "segment": "medical devices"},
            ],
        },
    )

    wjson(
        "vendor-list.json",
        {
            "generated": str(date.today()),
            "vendors": [
                {"id": "V001", "name": "Midwest Resin Co.", "category": "resin", "lead_time_days": 14},
                {"id": "V002", "name": "Columbus Packaging", "category": "boxes", "lead_time_days": 7},
            ],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-02-01",
            "kpis": [
                {"name": "Scrap (ops)", "definition": "scrap incl rework"},
                {"name": "Scrap (QA)", "definition": "true scrap only"},
                {"name": "OTD (customer)", "definition": "on-time if first shipment arrives"},
            ],
        },
    )

    wjson(
        "expedite-request-form.json",
        {
            "schema_version": "0.1",
            "title": "Expedite Request",
            "fields": [
                {"id": "customer", "type": "string", "required": True},
                {"id": "sku", "type": "string", "required": True},
                {"id": "qty", "type": "integer", "required": True},
                {"id": "need_by", "type": "date", "required": True},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "customer-scorecard-2025-q4.pdf",
        "Customer Scorecard — Q4 2025 (Summary)",
        [
            "OTD (customer definition): 95%",
            "Quality NCRs: 3",
            "Note: partial shipments treated as on-time if first shipment hits date.",
        ],
    )


def make_xlsx():
    import openpyxl

    # revenue history
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Quarterly"
    ws.append(["Year","Quarter","Revenue","Notes"])
    for r in [(2024,"Q4",3600000,""),(2025,"Q1",3400000,""),(2025,"Q2",3800000,""),(2025,"Q3",3900000,""),(2025,"Q4",4200000,"")]: ws.append(list(r))
    for i in range(2, ws.max_row+1): ws.cell(i,3).number_format="$#,##0"
    wb.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="2026"
    ws2.append(["Month","Revenue","Notes"])
    for r in [("2026-01",1200000,""),("2026-02",1150000,""),("2026-03",1250000,"")]: ws2.append(list(r))
    for i in range(2, ws2.max_row+1): ws2.cell(i,2).number_format="$#,##0"
    wb2.save(BASE/"projected-revenue-2026.xlsx")

    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="Inventory"
    ws3.append(["Item","Qty","Std Cost","Last Purchase","Valuation (finance)","Valuation (ops)"])
    ws3.append(["Resin (lb)",18000,0.88,0.94,None,None])
    ws3.cell(2,5).value="=B2*C2"; ws3.cell(2,6).value="=B2*D2"
    wb3.save(BASE/"inventory-valuation-2025-12.xlsx")

    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="H1 2026"
    ws4.append(["Week","Line","Primary SKU","Planned Qty","Notes"])
    for r in [("2026-W02","Line 1","TRAY-01",50000,""),("2026-W03","Line 2","INSERT-04",32000,"changeover")]: ws4.append(list(r))
    wb4.save(BASE/"production-plan-2026-h1.xlsx")

    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="Incidents 2025"
    ws5.append(["Date","Part","Issue","Severity","Notes"])
    for r in [("2025-11-18","TRAY-01","Warping","Med","rework"),("2025-12-09","INSERT-04","Short shots","High","resin moisture")]: ws5.append(list(r))
    wb5.save(BASE/"qa-incidents-log-2025.xlsx")

    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="2026-01"
    ws6.append(["Week","Ops scrap %","QA scrap %","Notes"])
    for r in [("2026-W01",0.042,0.031,"rework excluded by QA"),("2026-W02",0.039,0.029,"")]: ws6.append(list(r))
    for i in range(2, ws6.max_row+1):
        ws6.cell(i,2).number_format="0.0%"; ws6.cell(i,3).number_format="0.0%"
    wb6.save(BASE/"scrap-rework-tracker-2026-01.xlsx")


def make_pptx():
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"marketing-pitch-deck.pptx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx(); make_pptx()


if __name__ == "__main__":
    main()
