from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Alder Ridge Builders — Mission, Vision, Values

## Mission
Deliver high-quality remodels and small builds with predictable schedules and honest change management.

## Vision
Be the most trusted small GC in the Portland metro for ADUs and remodels.

## Values
- Safety and respect on site
- Clear documentation
- No surprises (or handle them fast)
- Do it right the first time
""",
    )

    w(
        "brand-guide.md",
        """# Alder Ridge Builders — Brand Guide (v0.1)

## Voice
- Friendly, direct, practical.

## Palette (suggested)
- Forest green: #1F3D2B
- Warm gray: #6B7280
- Off-white: #F4EFE6

## Visual style
- Before/after photos
- Simple diagrams for schedules and budgets
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Alex Rivera
Last updated: 2026-01-05

## Focus
- Keep jobs under control (schedule + budget)
- Standardize change orders
- Build repeatable subcontractor bench

## Targets
- 8–10 projects/year
- Gross margin target: 28–32%
""",
    )

    w(
        "project-plan-adu-cedar-st.md",
        """# Project Plan — ADU (Cedar St)

Client: Cedar Street ADU (job nickname)
Start: 2026-01-08

## Scope
- Foundation + framing
- Rough-in MEP
- Insulation + drywall
- Finishes

## Milestones
- Rough-in complete: Feb
- Drywall: Mar
- Substantial completion: Apr

## Risks
- Permit inspection availability
- Long lead windows (if upgraded)
""",
    )

    w(
        "meeting-notes-adu-cedar-st-2026-01-10.md",
        """# Meeting Notes — ADU (Cedar St) — 2026-01-10

Attendees: Alex, Marcus, client

## Decisions
- Confirm finish selections by Feb 1

## Actions
- Brooke to confirm inspection timeline
""",
    )

    w(
        "meeting-notes-adu-cedar-st-2026-02-07.md",
        """# Meeting Notes — ADU (Cedar St) — 2026-02-07

## Update
- Client requested upgraded windows (scope add)

## Next
- Issue change order CO-02
- Update schedule
""",
    )

    w(
        "subcontractor-evaluation-template.md",
        """# Subcontractor Evaluation Template

Sub:
Trade:

Scores (1–5):
- Quality:
- Schedule reliability:
- Communication:
- Safety:

Notes:
""",
    )

    w(
        "safety-walk-template.md",
        """# Safety Walk Template

Date:
Site:

Checks:
- PPE in use
- Housekeeping
- Fall protection
- Electrical safety
- Tool condition

Notes:
""",
    )


def make_json():
    wjson(
        "subcontractor-list.json",
        {
            "generated": str(date.today()),
            "subs": [
                {"name": "Rose City Electric", "trade": "Electrical"},
                {"name": "Bridgeview Plumbing", "trade": "Plumbing"},
                {"name": "Cascade HVAC", "trade": "HVAC"},
                {"name": "Northside Drywall", "trade": "Drywall"},
                {"name": "Pine & Paint", "trade": "Painting"},
            ],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-01-05",
            "kpis": [
                {"name": "Gross margin", "definition": "(Contract - costs) / contract"},
                {"name": "Schedule variance", "definition": "actual milestone date - planned milestone date"},
            ],
        },
    )

    wjson(
        "client-intake-form.json",
        {
            "schema_version": "0.1",
            "title": "Client Intake",
            "fields": [
                {"id": "client_name", "type": "string", "required": True},
                {"id": "project_address", "type": "string", "required": True},
                {"id": "project_type", "type": "enum", "values": ["ADU", "Kitchen", "Bath", "TI"], "required": True},
                {"id": "target_start", "type": "date"},
                {"id": "budget_range", "type": "string"},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "company-capabilities-one-pager.pdf",
        "Alder Ridge Builders — Capabilities",
        [
            "Residential remodels and ADUs",
            "Small commercial tenant improvements",
            "Clear estimating + change orders",
            "Licensed + insured (fictional)",
        ],
    )

    pdf(
        BASE / "sample-estimate-adu-cedar-st.pdf",
        "Estimate — ADU (Cedar St)",
        [
            "Base contract: $265,000",
            "Allowances: fixtures, finishes",
            "Exclusions: unforeseen structural repairs",
        ],
    )

    pdf(
        BASE / "contract-summary-adu-cedar-st.pdf",
        "Contract Summary — ADU (Cedar St)",
        [
            "Contract type: fixed bid",
            "Payment schedule: milestones",
            "Change orders: written approval required",
        ],
    )


def make_xlsx():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    # Estimate template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws.append(["Category", "Description", "Qty", "Unit", "Unit Cost", "Line Total"])
    items = [
        ("Demo", "Selective demolition", 1, "lot", 8500),
        ("Framing", "Framing labor", 1, "lot", 42000),
        ("MEP", "Rough-in MEP", 1, "lot", 38000),
        ("Finishes", "Drywall + paint", 1, "lot", 26000),
        ("Finishes", "Flooring", 1, "lot", 14500),
    ]
    for cat, desc, qty, unit, unit_cost in items:
        ws.append([cat, desc, qty, unit, unit_cost, None])

    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F3D2B")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    for r in range(2, ws.max_row + 1):
        qty = ws.cell(r, 3).coordinate
        unit_cost = ws.cell(r, 5).coordinate
        ws.cell(r, 6).value = f"={qty}*{unit_cost}"
        ws.cell(r, 5).number_format = "$#,##0"
        ws.cell(r, 6).number_format = "$#,##0"

    total_row = ws.max_row + 2
    ws.cell(total_row, 5).value = "TOTAL"
    ws.cell(total_row, 6).value = f"=SUM(F2:F{ws.max_row})"
    ws.cell(total_row, 6).number_format = "$#,##0"

    wb.save(BASE / "estimate-template.xlsx")

    # Change order log
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Change Orders"
    ws2.append(["CO #", "Date", "Description", "Cost", "Days", "Approved?", "Notes"])
    rows = [
        ("CO-01", "2026-01-20", "Upgrade insulation", 3200, 1, "Y", ""),
        ("CO-02", "2026-02-08", "Upgrade windows", 9800, 5, "Y", "Lead time risk"),
    ]
    for r in rows:
        ws2.append(list(r))
    for r in range(2, ws2.max_row + 1):
        ws2.cell(r, 4).number_format = "$#,##0"
    ws2.cell(ws2.max_row + 2, 3).value = "Total CO cost"
    ws2.cell(ws2.max_row, 4).value = f"=SUM(D2:D{ws2.max_row-2})"
    wb2.save(BASE / "change-order-log-adu-cedar-st.xlsx")

    # Schedule
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Schedule"
    ws3.append(["Task", "Start", "End", "Owner", "Notes"])
    tasks = [
        ("Foundation", "2026-01-08", "2026-01-22", "Marcus", ""),
        ("Framing", "2026-01-23", "2026-02-14", "Marcus", ""),
        ("Rough-in MEP", "2026-02-17", "2026-03-07", "Alex", "Subs coordinated"),
        ("Drywall", "2026-03-10", "2026-03-24", "Brooke", ""),
        ("Finishes", "2026-03-25", "2026-04-20", "Alex", ""),
    ]
    for t in tasks:
        ws3.append(list(t))
    wb3.save(BASE / "schedule-adu-cedar-st.xlsx")

    # RFI log
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = "RFIs"
    ws4.append(["RFI #", "Date", "Question", "Response", "Status"])
    rfis = [
        ("RFI-01", "2026-01-12", "Confirm window rough openings", "Use manufacturer spec", "Closed"),
        ("RFI-02", "2026-02-09", "Confirm upgraded window lead time", "4-6 weeks", "Closed"),
    ]
    for r in rfis:
        ws4.append(list(r))
    wb4.save(BASE / "rfi-log-adu-cedar-st.xlsx")

    # Revenue history + projection
    wb5 = openpyxl.Workbook()
    ws5 = wb5.active
    ws5.title = "Quarterly"
    ws5.append(["Year", "Quarter", "Revenue", "Notes"])
    for row in [(2024,"Q3",310000,""),(2024,"Q4",420000,""),(2025,"Q1",290000,""),(2025,"Q2",360000,""),(2025,"Q3",410000,""),(2025,"Q4",520000,"" )]:
        ws5.append(list(row))
    for r in range(2, ws5.max_row+1):
        ws5.cell(r,3).number_format = "$#,##0"
    wb5.save(BASE / "quarterly-revenue-history-2024-2025.xlsx")

    wb6 = openpyxl.Workbook()
    ws6 = wb6.active
    ws6.title = "2026"
    ws6.append(["Month","Revenue","Notes"])
    for row in [("2026-01",90000,""),("2026-02",85000,""),("2026-03",110000,"" )]:
        ws6.append(list(row))
    for r in range(2, ws6.max_row+1):
        ws6.cell(r,2).number_format = "$#,##0"
    wb6.save(BASE / "projected-revenue-2026.xlsx")

    # Job costing
    wb7 = openpyxl.Workbook()
    ws7 = wb7.active
    ws7.title = "Job Cost"
    ws7.append(["Cost Code","Description","Budget","Actual to date","Variance"])
    rows = [
        ("01-100","Demo",8500,7800,None),
        ("03-200","Framing",42000,39800,None),
        ("05-300","MEP",38000,12000,None),
        ("09-400","Finishes",55000,6000,None),
    ]
    for r in rows:
        ws7.append(list(r))
    for r in range(2, ws7.max_row+1):
        bud = ws7.cell(r,3).coordinate
        act = ws7.cell(r,4).coordinate
        ws7.cell(r,5).value = f"={bud}-{act}"
        for c in [3,4,5]:
            ws7.cell(r,c).number_format = "$#,##0"
    wb7.save(BASE / "job-costing-adu-cedar-st.xlsx")

    # AR
    wb8 = openpyxl.Workbook()
    ws8 = wb8.active
    ws8.title = "AR"
    ws8.append(["Client","Invoice","Invoice Date","Amount","Received Date","Status"])
    rows = [
        ("Cedar St ADU","INV-1042","2026-02-05",38000,"2026-02-12","Paid"),
        ("Hawthorne Kitchen","INV-1037","2026-01-28",22000,"","Open"),
    ]
    for r in rows:
        ws8.append(list(r))
    for r in range(2, ws8.max_row+1):
        ws8.cell(r,4).number_format = "$#,##0"
    wb8.save(BASE / "accounts-receivable-2026-02.xlsx")


def main():
    make_markdown()
    make_json()
    make_pdfs()
    make_xlsx()


if __name__ == "__main__":
    main()
