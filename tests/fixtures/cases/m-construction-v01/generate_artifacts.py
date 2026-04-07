from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Beaconline Construction Co. — Mission, Vision, Values

## Mission
Deliver tenant improvements and light industrial projects with reliable schedules, clean change management, and job-cost transparency.

## Vision
Be the mid-sized GC clients trust to execute without drama.

## Values
- Safety first
- Clear documentation
- Own the schedule
- Be fair to subs
""",
    )

    w(
        "brand-guide.md",
        """# Beaconline Construction Co. — Brand Guide (v0.1)

## Voice
- Professional, direct, schedule/budget anchored.

## Palette
- Navy: #0B1F3A
- Safety yellow: #FBBF24
- Off-white: #F7F5F0
- Gray: #6B7280

## Visual style
- Site progress photos
- Simple diagrams (schedule, CO flow, WIP)
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Kim Alvarez
Last updated: 2026-01-03

## Focus
- Reduce schedule slippage via long-lead tracking
- Tighten change order discipline (approved vs pending)
- Align PM cost reports with accounting WIP

## Targets
- 20–30 projects/year
- Maintain fee margin in 4–6% range
""",
    )

    w(
        "project-plan-retail-ti-pine-st.md",
        """# Project Plan — Retail TI (Pine St)

Project: Retail TI — Pine St
Start: 2026-01-06

## Scope
- Demo + framing
- MEP rough-in
- Storefront glass + doors
- Finishes

## Risks
- Long lead storefront glass
- Night work constraints

## Baseline milestones
- Rough-in complete: Feb
- Substantial completion: late Mar (may slip)
""",
    )

    w(
        "meeting-notes-retail-ti-pine-st-2026-01-09.md",
        """# Weekly Meeting — Retail TI (Pine St) — 2026-01-09

Attendees: Sara (PM), Devon (Super), subs

## Notes
- Confirm demo start
- Submittals due next week

## Actions
- Devon to confirm glass lead time
""",
    )

    w(
        "meeting-notes-retail-ti-pine-st-2026-02-06.md",
        """# Weekly Meeting — Retail TI (Pine St) — 2026-02-06

## Schedule
- Substantial completion still targeted: **2026-03-28** (per meeting notes)

## Change orders
- CO-03 pending (storefront upgrade)
""",
    )

    w(
        "meeting-notes-retail-ti-pine-st-2026-03-06.md",
        """# Weekly Meeting — Retail TI (Pine St) — 2026-03-06

## Schedule
- Re-baseline after glass delay.

## Notes
- Team agrees updated schedule will reflect new substantial completion date.
""",
    )

    w(
        "owner-summary-email-2026-03-07.md",
        """# Owner Summary Email — 2026-03-07 (Kim)

Subject: Pine St TI — status

- Schedule: re-baselined; substantial completion early April.
- Change orders: total **$124,900** including pending items.
- Margin: watching; avoid further scope creep.
""",
    )

    w(
        "subcontractor-evaluation-template.md",
        """# Subcontractor Evaluation Template

Sub:
Trade:

Scores (1–5):
- Quality
- Schedule reliability
- Communication
- Safety

Notes:
""",
    )

    w(
        "safety-walk-template.md",
        """# Safety Walk Template

Date:
Site:

Checks:
- PPE
- Housekeeping
- Fall protection
- Hot work controls
- Tool condition

Notes:
""",
    )

    w(
        "incident-report-template.md",
        """# Incident Report Template

Date:
Project:
Description:
Immediate actions:
Root cause:
Corrective actions:
Owner:
""",
    )


def make_json():
    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-01-03",
            "kpis": [
                {"name": "Fee margin % (PM)", "definition": "(contract - costs) / contract"},
                {"name": "Fee margin % (WIP)", "definition": "(revenue recognized - costs) / revenue recognized"},
            ],
        },
    )

    wjson(
        "client-intake-form.json",
        {
            "schema_version": "0.1",
            "title": "Client Intake",
            "fields": [
                {"id": "client_legal_name", "type": "string", "required": True},
                {"id": "project_name", "type": "string", "required": True},
                {"id": "site_address", "type": "string", "required": True},
                {"id": "project_type", "type": "enum", "values": ["TI", "Industrial", "Common Area"], "required": True},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "company-capabilities-one-pager.pdf",
        "Beaconline Construction — Capabilities",
        [
            "Tenant improvements (office/retail)",
            "Light industrial buildouts",
            "Long-lead tracking + change order discipline",
        ],
    )

    pdf(
        BASE / "sample-estimate-retail-ti-pine-st.pdf",
        "Estimate — Retail TI (Pine St)",
        [
            "Base contract: $1,180,000",
            "Contingency (GMP): $45,000",
            "Exclusions: landlord scope items",
        ],
    )

    pdf(
        BASE / "contract-summary-retail-ti-pine-st.pdf",
        "Contract Summary — Retail TI (Pine St)",
        [
            "Contract type: GMP",
            "Change orders: approved vs pending tracked separately",
            "Schedule: baseline + re-baselines",
        ],
    )


def make_xlsx():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    # Estimate template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws.append(["Division", "Description", "Qty", "Unit", "Unit Cost", "Line Total"])
    items = [
        ("01", "General conditions", 1, "lot", 120000),
        ("02", "Demo", 1, "lot", 85000),
        ("06", "Rough carpentry", 1, "lot", 210000),
        ("22", "Plumbing", 1, "lot", 140000),
        ("23", "HVAC", 1, "lot", 155000),
        ("26", "Electrical", 1, "lot", 185000),
        ("08", "Doors/frames/hardware", 1, "lot", 98000),
        ("09", "Finishes", 1, "lot", 260000),
    ]
    for div, desc, qty, unit, unit_cost in items:
        ws.append([div, desc, qty, unit, unit_cost, None])

    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="0B1F3A")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    for r in range(2, ws.max_row + 1):
        qty = ws.cell(r, 3).coordinate
        unit_cost = ws.cell(r, 5).coordinate
        ws.cell(r, 6).value = f"={qty}*{unit_cost}"
        ws.cell(r, 5).number_format = "$#,##0"
        ws.cell(r, 6).number_format = "$#,##0"

    total_row = ws.max_row + 2
    ws.cell(total_row, 5).value = "TOTAL"
    ws.cell(total_row, 6).value = f"=SUM(F2:F{ws.max_row})"
    ws.cell(total_row, 6).number_format = "$#,##0"

    wb.save(BASE / "estimate-template.xlsx")

    # Schedule with re-baseline contradiction
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Schedule"
    ws2.append(["Task", "Start", "End", "Owner", "Milestone?"])
    tasks = [
        ("Demo", "2026-01-10", "2026-01-24", "Devon", ""),
        ("Rough-in MEP", "2026-01-27", "2026-02-21", "Sara", ""),
        ("Storefront glass install", "2026-02-24", "2026-03-20", "Devon", ""),
        ("Finishes", "2026-03-10", "2026-04-02", "Sara", ""),
        ("Substantial Completion", "", "2026-04-05", "Sara", "Y"),
    ]
    for t in tasks:
        ws2.append(list(t))
    wb2.save(BASE / "schedule-retail-ti-pine-st.xlsx")

    # Logs
    def log_book(path, sheet, headers, rows, money_cols=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(headers)
        for r in rows:
            ws.append(list(r))
        if money_cols:
            for rr in range(2, ws.max_row + 1):
                for c in money_cols:
                    ws.cell(rr, c).number_format = "$#,##0"
        wb.save(BASE / path)

    log_book(
        "rfi-log-retail-ti-pine-st.xlsx",
        "RFIs",
        ["RFI #", "Date", "Question", "Response", "Status", "Owner"],
        [
            ("RFI-01", "2026-01-15", "Confirm demo boundaries", "Approved", "Closed", "Sara"),
            ("RFI-07", "2026-02-18", "Storefront glass spec", "Pending", "Open", "Devon"),
        ],
    )

    log_book(
        "submittal-log-retail-ti-pine-st.xlsx",
        "Submittals",
        ["Submittal", "Date", "Spec", "Status", "Owner"],
        [
            ("Glass shop drawings", "2026-01-22", "08 44 13", "Revise", "Devon"),
            ("Lighting package", "2026-01-28", "26 51 00", "Approved", "Sara"),
        ],
    )

    log_book(
        "change-order-log-retail-ti-pine-st.xlsx",
        "Change Orders",
        ["CO #", "Date", "Description", "Cost", "Days", "Status"],
        [
            ("CO-01", "2026-01-30", "After-hours demo", 18400, 2, "Approved"),
            ("CO-02", "2026-02-12", "Electrical scope add", 36000, 1, "Approved"),
            ("CO-03", "2026-02-22", "Storefront upgrade", 64000, 8, "Pending"),
        ],
        money_cols=[4],
    )

    # Job costing + WIP contradiction
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Job Cost"
    ws3.append(["Cost Code", "Budget", "Committed", "Actual", "Forecast to Complete", "Variance"])
    rows = [
        ("GC", 120000, 115000, 82000, 122000, None),
        ("Demo", 85000, 91000, 74000, 94000, None),
        ("MEP", 480000, 510000, 220000, 525000, None),
        ("Finishes", 260000, 255000, 90000, 270000, None),
    ]
    for r in rows:
        ws3.append(list(r))
    for r in range(2, ws3.max_row + 1):
        bud = ws3.cell(r, 2).coordinate
        fc = ws3.cell(r, 5).coordinate
        ws3.cell(r, 6).value = f"={bud}-{fc}"
        for c in [2, 3, 4, 5, 6]:
            ws3.cell(r, c).number_format = "$#,##0"
    wb3.save(BASE / "job-costing-retail-ti-pine-st.xlsx")

    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = "WIP"
    ws4.append(["Project", "Contract", "CO Approved", "Cost to Date", "Est Cost at Comp", "Margin % (WIP)"])
    ws4.append(["Retail TI — Pine St", 1180000, 54400, 610000, 1208000, None])
    ws4.cell(2, 2).number_format = "$#,##0"
    ws4.cell(2, 3).number_format = "$#,##0"
    ws4.cell(2, 4).number_format = "$#,##0"
    ws4.cell(2, 5).number_format = "$#,##0"
    # margin% = (contract+co - est cost)/ (contract+co)
    ws4.cell(2, 6).value = "=(B2+C2-E2)/(B2+C2)"
    ws4.cell(2, 6).number_format = "0.0%"
    wb4.save(BASE / "wip-report-2026-02.xlsx")

    # Revenue history + projection
    wb5 = openpyxl.Workbook()
    ws5 = wb5.active
    ws5.title = "Quarterly"
    ws5.append(["Year", "Quarter", "Revenue", "Notes"])
    for row in [(2024,"Q3",1900000,""),(2024,"Q4",2400000,""),(2025,"Q1",2100000,""),(2025,"Q2",2600000,""),(2025,"Q3",2550000,""),(2025,"Q4",3100000,"")]:
        ws5.append(list(row))
    for r in range(2, ws5.max_row + 1):
        ws5.cell(r, 3).number_format = "$#,##0"
    wb5.save(BASE / "quarterly-revenue-history-2024-2025.xlsx")

    wb6 = openpyxl.Workbook()
    ws6 = wb6.active
    ws6.title = "2026"
    ws6.append(["Month", "Revenue", "Notes"])
    for row in [("2026-01",580000,""),("2026-02",620000,""),("2026-03",700000,"" )]:
        ws6.append(list(row))
    for r in range(2, ws6.max_row + 1):
        ws6.cell(r, 2).number_format = "$#,##0"
    wb6.save(BASE / "projected-revenue-2026.xlsx")

    # AR
    wb7 = openpyxl.Workbook()
    ws7 = wb7.active
    ws7.title = "AR"
    ws7.append(["Client", "Invoice", "Invoice Date", "Amount", "Received Date", "Status"])
    ws7.append(["Pine St Retail", "INV-2201", "2026-02-04", 185000, "2026-02-28", "Paid"])
    ws7.append(["Ballard Office TI", "INV-2192", "2026-01-26", 97000, "", "Open"])
    for r in range(2, ws7.max_row + 1):
        ws7.cell(r, 4).number_format = "$#,##0"
    wb7.save(BASE / "accounts-receivable-2026-02.xlsx")


def main():
    make_markdown()
    make_json()
    make_pdfs()
    make_xlsx()


if __name__ == "__main__":
    main()
