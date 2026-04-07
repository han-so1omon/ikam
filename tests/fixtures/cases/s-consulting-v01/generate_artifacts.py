from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_md():
    w(
        "mission-vision-values.md",
        """# Birchline Insights — Mission, Vision, Values

Mission: make complex ops decisions measurable.
Vision: a boutique partner known for clarity and credible analysis.
Values: transparency, pragmatism, good definitions.
""",
    )

    w(
        "brand-guide.md",
        """# Birchline Insights — Brand Guide (v0.1)

Palette:
- Forest: #14532D
- Slate: #334155
- Cream: #FAF7F0

Voice: concise, analytical.
""",
    )

    w(
        "change-request-001.md",
        """# Change Request 001 — Redwood Freight

Requested: 2026-01-27

Summary: Add light implementation support after discovery.

Notes:
- This conflicts with the original SOW scope if not formally executed.
""",
    )

    w(
        "project-plan.md",
        """# Project Plan — Redwood Freight Network Optimization (Discovery)

Duration: 6 weeks
Workstreams:
- Data intake
- Baseline network model
- Scenario evaluation

Open question: implementation support requested late in discovery.
""",
    )

    for d, txt in [
        ("weekly-status-2026-01-12.md", "Week 1: discovery kickoff; data intake pending."),
        ("weekly-status-2026-01-19.md", "Week 2: lane data cleaned; early hypotheses."),
        ("weekly-status-2026-01-26.md", "Week 3: scenario modeling; client asked about implementation.")
    ]:
        w(d, f"# Weekly Status — {d[-14:-3]}\n\n{txt}\n")

    w(
        "client-email-thread-2026-01-27.md",
        """# Client Email Thread — 2026-01-27

Client: "Can you also help implement the routing changes?"
Birchline: "We can, but the current SOW is discovery only. We’ll send a change request."

Note: deck labeled FINAL was still described as draft pending late-arriving client data.
""",
    )

    w(
        "sow-template.md",
        """# SOW Template

Scope:
Deliverables:
Timeline:
Assumptions:
Out of scope:
Fees:
""",
    )

    w(
        "invoice-template.md",
        """# Invoice Template

Invoice #:
Period:
Hours:
Rate:
Write-offs:
Total:
""",
    )

    w(
        "meeting-notes-template.md",
        """# Meeting Notes Template

Attendees:
Topics:
Decisions:
Actions:
""",
    )


def make_json():
    wjson(
        "org.json",
        {
            "generated": str(date.today()),
            "team": [
                {"name": "Avery Hart", "role": "Principal"},
                {"name": "Jordan Ng", "role": "Consultant"},
                {"name": "Casey Moore", "role": "Analyst"},
                {"name": "Priya Desai", "role": "Fractional Finance"},
            ],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-01-31",
            "kpis": [
                {"name": "Billable hours", "definition": "hours invoiced after write-offs"},
                {"name": "Delivered hours", "definition": "timesheet hours logged"},
                {"name": "Write-off", "definition": "hours removed from invoice"},
            ],
        },
    )


def make_xlsx():
    import openpyxl

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Timesheet"
    ws.append(["Date","Person","Client","Project","Hours","Notes"])
    rows=[("2026-01-12","Avery","Redwood Freight","Discovery",6,""),
          ("2026-01-13","Jordan","Redwood Freight","Discovery",8,""),
          ("2026-01-14","Casey","Redwood Freight","Discovery",7,""),
          ("2026-01-20","Jordan","Redwood Freight","Discovery",9,""),
          ("2026-01-21","Casey","Redwood Freight","Discovery",8,""),
          ("2026-01-26","Avery","Redwood Freight","Discovery",6,""),
          ("2026-01-27","Jordan","Redwood Freight","Discovery",7,""),
          ("2026-01-28","Casey","Redwood Freight","Discovery",6,""),
          ]
    for r in rows: ws.append(list(r))
    ws.append(["","","","TOTAL",None,""])
    ws.cell(ws.max_row,5).value=f"=SUM(E2:E{ws.max_row-1})"
    wb.save(BASE/"timesheet-2026-01.xlsx")

    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Pipeline"
    ws2.append(["Client","Stage","Close","Value","Probability","Weighted"])
    ws2.append(["Mesa Dental Group","Proposal","2026-03-15",65000,0.45,None])
    ws2.append(["Skyline Foods","Discovery","2026-04-10",90000,0.30,None])
    for i in range(2, ws2.max_row+1):
        ws2.cell(i,6).value=f"=D{i}*E{i}"
        ws2.cell(i,4).number_format="$#,##0"; ws2.cell(i,6).number_format="$#,##0"; ws2.cell(i,5).number_format="0%"
    wb2.save(BASE/"pipeline.xlsx")

    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="2025"
    ws3.append(["Month","Revenue"])
    for r in [("2025-09",12000),("2025-10",18500),("2025-11",24000),("2025-12",26500)]:
        ws3.append(list(r)); ws3.cell(ws3.max_row,2).number_format="$#,##0"
    wb3.save(BASE/"revenue-history-2025.xlsx")


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "services-one-pager.pdf",
        "Birchline Insights — Services",
        ["Operations diagnostics", "Network optimization", "KPI design", "Light implementation support"],
    )

    pdf(
        BASE / "sow-redwood-freight.pdf",
        "Statement of Work — Redwood Freight (Discovery)",
        [
            "Scope: 6-week discovery only.",
            "Deliverable: recommendations deck.",
            "Out of scope: implementation unless via change request.",
        ],
    )

    # invoice: mismatch vs timesheet by including write-offs
    pdf(
        BASE / "invoice-summary-2026-01.pdf",
        "Invoice Summary — 2026-01",
        [
            "Timesheet hours: 148 (internal tally)",
            "Billed hours: 132 (includes 16 hours write-off)",
            "Notes: client requested extra analysis; goodwill write-off applied.",
        ],
    )


def make_pptx():
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"deliverable-deck-final.pptx")
    prs=Presentation(); prs.save(BASE/"deliverable-deck-v2.pptx")


def main():
    make_md(); make_json(); make_xlsx(); make_pdfs(); make_pptx()


if __name__ == "__main__":
    main()
