from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Ridgeway Metal Works — Mission, Vision, Values

Mission: Make reliable parts fast with honest communication.
Vision: Be the go-to job shop for repeat customers.
Values: Safety, quality, keep promises.
""",
    )

    w(
        "brand-guide.md",
        """# Ridgeway Metal Works — Brand Guide (v0.1)

Palette:
- Steel gray: #2F3A44
- Orange: #F97316

Voice: practical, direct.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Hank Miller

- Keep lead times tight
- Improve quoting accuracy
- Track material price swings
""",
    )

    w(
        "supplier-evaluation-template.md",
        """# Supplier Evaluation Template

Supplier:
Scores: lead time, quality, pricing.
Notes:
""",
    )

    w(
        "nonconformance-report-template.md",
        """# Nonconformance Report Template

Date:
Job:
Issue:
Containment:
Root cause:
Fix:
""",
    )


def make_json():
    wjson(
        "customer-list.json",
        {
            "generated": str(date.today()),
            "customers": [
                {"account": "Spokane HVAC Supply", "segment": "distributor"},
                {"account": "Pine River Equipment", "segment": "OEM"},
            ],
        },
    )

    wjson(
        "vendor-list.json",
        {
            "generated": str(date.today()),
            "vendors": [
                {"id": "V001", "name": "Inland Steel", "category": "steel", "lead_time_days": 7},
                {"id": "V002", "name": "Orange County Fasteners", "category": "hardware", "lead_time_days": 5},
            ],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-01-10",
            "kpis": [
                {"name": "On-time", "definition": "jobs shipped by promised date"},
                {"name": "Job margin", "definition": "(invoice - costs) / invoice"},
            ],
        },
    )

    wjson(
        "job-board-snapshot-2026-02.json",
        {
            "snapshot_date": "2026-02-06",
            "board": {
                "Queued": ["J-1042", "J-1046"],
                "Cutting": ["J-1039"],
                "Welding": ["J-1040"],
                "Done": ["J-1034"],
            },
        },
    )

    wjson(
        "customer-intake-form.json",
        {
            "schema_version": "0.1",
            "title": "Customer Intake",
            "fields": [
                {"id": "company", "type": "string", "required": True},
                {"id": "contact_email", "type": "string"},
                {"id": "part_description", "type": "string", "required": True},
                {"id": "quantity", "type": "integer", "required": True},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "capabilities-one-pager.pdf",
        "Ridgeway Metal Works — Capabilities",
        ["Laser cutting", "Welding", "Short-run brackets/plates", "Fast turnaround"],
    )

    pdf(
        BASE / "sample-quote-bracket-run.pdf",
        "Sample Quote — Bracket Run",
        ["Qty: 500", "Material: mild steel", "Lead time: 10 business days", "Note: material pricing subject to change"],
    )


def make_xlsx():
    import openpyxl

    # Quote template
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Quote"
    ws.append(["Line","Description","Qty","Unit","Unit Cost","Line Total"])
    ws.append([1,"Laser cut bracket",500,"ea",4.20,None])
    ws.cell(2,6).value="=C2*E2"; ws.cell(2,6).number_format="$#,##0.00"; ws.cell(2,5).number_format="$0.00"
    ws.append(["","","","","TOTAL",None])
    ws.cell(3,6).value="=SUM(F2:F2)"; ws.cell(3,6).number_format="$#,##0.00"
    wb.save(BASE/"quote-template.xlsx")

    # Production schedule
    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="2026-02"
    ws2.append(["Job","Stage","Due","Owner","Notes"])
    for r in [("J-1039","Cutting","2026-02-10","Eli",""),("J-1040","Welding","2026-02-12","Eli",""),("J-1042","Queued","2026-02-18","Hank","repeat customer")]:
        ws2.append(list(r))
    wb2.save(BASE/"production-schedule-2026-02.xlsx")

    # Inventory count
    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="Count"
    ws3.append(["Item","Qty","Unit","Notes"])
    for r in [("Steel sheet 4x8",12,"sheets","manual count"),("Welding wire",6,"spools","")]:
        ws3.append(list(r))
    wb3.save(BASE/"inventory-count-sheet.xlsx")

    # Incident log
    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="Q1 2026"
    ws4.append(["Date","Incident","Severity","Notes"])
    for r in [("2026-01-22","Laser cutter down 3h","Med",""),("2026-02-05","Material late delivery","Low","")]:
        ws4.append(list(r))
    wb4.save(BASE/"incident-log-2026-q1.xlsx")

    # Revenue history
    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="Quarterly"
    ws5.append(["Year","Quarter","Revenue","Notes"])
    for r in [(2024,"Q4",420000,""),(2025,"Q1",390000,""),(2025,"Q2",440000,""),(2025,"Q3",460000,""),(2025,"Q4",510000,"steel volatility")]:
        ws5.append(list(r))
    for i in range(2, ws5.max_row+1): ws5.cell(i,3).number_format="$#,##0"
    wb5.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="2026"
    ws6.append(["Month","Revenue","Notes"])
    for r in [("2026-01",140000,""),("2026-02",135000,""),("2026-03",155000,"")]:
        ws6.append(list(r))
    for i in range(2, ws6.max_row+1): ws6.cell(i,2).number_format="$#,##0"
    wb6.save(BASE/"projected-revenue-2026.xlsx")

    # Job costing sample
    wb7=openpyxl.Workbook(); ws7=wb7.active; ws7.title="Job Cost"
    ws7.append(["Job","Invoice","Material","Labor","Margin %"])
    ws7.append(["J-1039",5200,2100,1800,None])
    ws7.cell(2,5).value="=(B2-(C2+D2))/B2"; ws7.cell(2,5).number_format="0.0%"
    for c in [2,3,4]: ws7.cell(2,c).number_format="$#,##0"
    wb7.save(BASE/"job-costing-sample.xlsx")


def make_pptx():
    # keep minimal for small shop
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"marketing-pitch-deck.pptx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx(); make_pptx()


if __name__ == "__main__":
    main()
