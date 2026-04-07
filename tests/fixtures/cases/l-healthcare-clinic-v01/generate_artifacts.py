from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Harborview Care Partners — Mission, Vision, Values

Mission: accessible care with consistent standards across sites.
Vision: trusted neighborhood clinics with reliable operations.
Values: patient dignity, safety, clarity.
""",
    )

    w(
        "brand-guide.md",
        """# HCP — Brand Guide (v0.1)

Palette:
- Sea blue: #0EA5E9
- Slate: #334155
- White: #F8FAFC

Voice: calm, reassuring, precise.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Ethan Park

- Unify KPI definitions across ops + rev cycle
- Reduce denials + no-shows
- Standardize triage SOP rollout
""",
    )

    w(
        "clinical-sop-triage-v2.md",
        """# Clinical SOP — Triage (v2)

Effective: 2025-03-15

- Urgent symptoms escalated to provider within 10 minutes.
- Wait time metric used locally: door-to-doc.
""",
    )

    w(
        "clinical-sop-triage-v3-DRAFT.md",
        """# Clinical SOP — Triage (v3 DRAFT)

Drafted: 2026-01-28

Changes:
- Adds urgent care flow and defines wait time as arrival-to-departure.
- Clarifies late-cancel handling.

Known issue: not yet rolled out at all sites.
""",
    )

    w(
        "clinic-site-operations-playbook.md",
        """# Clinic Site Operations Playbook

- Front desk check-in
- Rooming
- Provider handoff
- Checkout

Note: Some sites use local scheduling templates.
""",
    )

    w(
        "call-center-script.md",
        """# Call Center Script (v1)

- Verify name + DOB (do not write down over the phone).
- Offer next available.
- Remind about cancellation policy.
""",
    )

    w(
        "compliance-audit-memo-2026-01.md",
        """# Compliance Audit Memo — 2026-01

Finding: HIPAA training completion cited at 91% based on updated roster.
Note: HR tracker shows higher completion due to an older employee list.
""",
    )

    w(
        "employee-handbook-v4.md",
        """# Employee Handbook (v4)

Effective: 2024-09-01
- Attendance policy references legacy no-show definition.
""",
    )

    w(
        "employee-handbook-v5.md",
        """# Employee Handbook (v5)

Effective: 2026-02-01
- Attendance policy updated; late-cancels treated as no-shows for internal ops.
""",
    )

    w(
        "ops-review-notes-2026-01-22.md",
        """# Ops Review Notes — 2026-01-22

- No-show report excludes late-cancels (ops dashboard).
- Rev cycle report includes late-cancels.
""",
    )

    w(
        "nursing-huddle-notes-2026-02-04.md",
        """# Nursing Huddle — 2026-02-04

- Triage SOP v3 draft circulating.
- Clinics still measuring door-to-doc.
""",
    )

    w(
        "voice-note-transcript-2026-02-10.md",
        """# Voice Note Transcript — 2026-02-10 (Sofia)

- Denials rising; payer policy changes.
- KPI definitions are not aligned.
""",
    )

    w(
        "incident-report-template.md",
        """# Incident Report Template

Date:
Site:
Type (incident/near-miss):
Description:
Actions:
Owner:
""",
    )

    w(
        "patient-complaint-template.md",
        """# Patient Complaint Template

Date:
Site:
Complaint:
Resolution:
Follow-up:
""",
    )

    w(
        "provider-onboarding-checklist.md",
        """# Provider Onboarding Checklist

- Credentialing complete
- HIPAA training
- EHR access
- Clinical SOP review
""",
    )


def make_json():
    wjson(
        "org-structure.json",
        {
            "generated": str(date.today()),
            "leaders": {
                "MedicalDirector": "Dr. Lila Moreno",
                "COO": "Ethan Park",
                "Nursing": "Naomi Chen",
                "RevCycle": "Sofia Alvarez",
                "Compliance": "Jordan Blake",
                "HR": "Priya Desai",
            },
            "sites": ["Clinic A", "Clinic B", "Clinic C", "Clinic D", "Clinic E", "Clinic F"],
        },
    )

    wjson(
        "patient-intake-form.json",
        {
            "schema_version": "0.1",
            "title": "Patient Intake",
            "fields": [
                {"id": "preferred_name", "type": "string"},
                {"id": "dob", "type": "date", "required": True},
                {"id": "insurance", "type": "string"},
                {"id": "reason_for_visit", "type": "string", "required": True},
            ],
        },
    )

    wjson(
        "billing-kpi-definitions.json",
        {
            "version": "2026-01-15",
            "kpis": [
                {"name": "No-show (rev cycle)", "definition": "includes late-cancels"},
                {"name": "No-show (ops dashboard)", "definition": "excludes late-cancels"},
                {"name": "Wait time (Clinic A)", "definition": "door-to-doc"},
                {"name": "Wait time (exec)", "definition": "arrival-to-departure"},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "urgent-care-protocols.pdf",
        "Urgent Care Protocols (Summary)",
        [
            "Triage escalation and red-flag symptoms",
            "Basic infection control reminders",
            "Documentation expectations",
        ],
    )

    pdf(
        BASE / "exec-scorecard-2026-01.pdf",
        "Exec Scorecard — Jan 2026 (Summary)",
        [
            "No-show rate: 6.1% (includes late cancels per rev cycle)",
            "Avg wait time: 54 min (arrival-to-departure)",
            "HIPAA training completion: 91% (audit roster)",
        ],
    )


def make_xlsx():
    import openpyxl

    # scheduling templates
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Clinic A"
    ws.append(["Date","Provider","Slots","Notes"])
    ws.append(["2026-02-10","Moreno",18,"door-to-doc metric tracked locally"])
    wb.save(BASE/"scheduling-template-clinic-a.xlsx")

    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Clinic D"
    ws2.append(["Date","Provider","Slots","Notes"])
    ws2.append(["2026-02-10","Locum",22,"uses arrival-to-departure"])
    wb2.save(BASE/"scheduling-template-clinic-d.xlsx")

    # HIPAA training tracker (shows 98%)
    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="Tracker"
    ws3.append(["Employee","Role","Site","Completed","Completion Date"])
    for r in [("E-102","MA","Clinic A","Yes","2026-01-05"),("E-214","RN","Clinic D","Yes","2026-01-12"),("E-301","FD","Clinic B","No","" )]:
        ws3.append(list(r))
    wb3.save(BASE/"hipaa-training-tracker-2026.xlsx")

    # denial log
    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="2025 Q4"
    ws4.append(["Date","Payer","Reason","Count","Notes"])
    ws4.append(["2025-11-15","PayerOne","medical necessity",42,"policy change"])
    wb4.save(BASE/"denial-log-2025-q4.xlsx")

    # no-show report contradiction
    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="2026-01"
    ws5.append(["Clinic","No-show (ops)","No-show (rev cycle)","Notes"])
    ws5.append(["Clinic A",0.045,0.061,"late cancels differ"])
    ws5.cell(2,2).number_format="0.0%"; ws5.cell(2,3).number_format="0.0%"
    wb5.save(BASE/"no-show-report-2026-01.xlsx")

    # payer mix
    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="2025"
    ws6.append(["Payer","Share"])
    for r in [("Commercial",0.56),("Medicare",0.22),("Medicaid",0.18),("Self-pay",0.04)]:
        ws6.append(list(r)); ws6.cell(ws6.max_row,2).number_format="0%"
    wb6.save(BASE/"payer-mix-2025.xlsx")

    # finance spreadsheets
    wb7=openpyxl.Workbook(); ws7=wb7.active; ws7.title="Quarterly"
    ws7.append(["Year","Quarter","Revenue"])
    for r in [(2024,"Q4",18200000),(2025,"Q1",17600000),(2025,"Q2",19100000),(2025,"Q3",20400000),(2025,"Q4",21200000)]:
        ws7.append(list(r))
    for i in range(2, ws7.max_row+1): ws7.cell(i,3).number_format="$#,##0"
    wb7.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb8=openpyxl.Workbook(); ws8=wb8.active; ws8.title="2026"
    ws8.append(["Month","Revenue"])
    for r in [("2026-01",7100000),("2026-02",6950000),("2026-03",7350000)]:
        ws8.append(list(r))
    for i in range(2, ws8.max_row+1): ws8.cell(i,2).number_format="$#,##0"
    wb8.save(BASE/"projected-revenue-2026.xlsx")

    wb9=openpyxl.Workbook(); ws9=wb9.active; ws9.title="2025"
    ws9.append(["Site","Revenue","Labor","Supplies","Rent","Net"])
    ws9.append(["Clinic A",5200000,2800000,620000,540000,None])
    ws9.cell(2,6).value="=B2-(C2+D2+E2)"
    for c in [2,3,4,5,6]: ws9.cell(2,c).number_format="$#,##0"
    wb9.save(BASE/"clinic-pnl-2025.xlsx")


def make_pptx():
    from pptx import Presentation
    prs=Presentation(); prs.save(BASE/"network-update-2026-02.pptx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx(); make_pptx()


if __name__ == "__main__":
    main()
