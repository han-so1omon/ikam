from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_markdown():
    w(
        "mission-vision-values.md",
        """# Ironwood Ridge Constructors — Mission, Vision, Values

Mission: deliver safe, predictable builds with transparent change management.
Vision: the contractor owners trust for complex schedules.
Values: safety, documentation, accountability.
""",
    )

    w(
        "brand-guide.md",
        """# Ironwood Ridge — Brand Guide (v0.1)

Palette:
- Charcoal: #1F2937
- Desert sand: #D6B47A
- Teal accent: #14B8A6

Voice: direct, professional.
""",
    )

    w(
        "high-level-strategy-2026.md",
        """# High-Level Strategy — 2026

Owner: Mateo Reyes

- One schedule source of truth per project
- CO clarity: approved vs pending vs forecast
- Align PM forecast with WIP
""",
    )

    w(
        "project-plan-desert-vista.md",
        """# Project Plan — Desert Vista Medical Office (DV-3291)

Scope: medical office + tenant improvements.

Key risks:
- MEP long lead
- inspections/weather

Target completion requested by owner: mid-Oct.
""",
    )

    w(
        "meeting-notes-desert-vista-2026-09-06.md",
        """# Weekly Meeting Notes — Desert Vista — 2026-09-06

- Long-lead MEP equipment tracking in a separate sheet.
- COs: 2 pending.
""",
    )

    w(
        "meeting-notes-desert-vista-2026-10-04.md",
        """# Weekly Meeting Notes — Desert Vista — 2026-10-04

- Internal schedule shows substantial completion early Nov.
- Owner still referencing Oct target.
""",
    )

    w(
        "owner-update-email-2026-09-12.md",
        """# Owner Update Email — 2026-09-12

Subject: Desert Vista — schedule + changes

- Substantial completion target: 2026-10-18
- CO total: "around" $1.0M including pending
""",
    )

    w(
        "exec-summary-email-2026-10-05.md",
        """# Exec Summary Email — 2026-10-05

- Substantial completion: 2026-11-05 (internal schedule)
- COs: $1.05M including pending
- WIP margin lower than PM forecast
""",
    )

    w(
        "safety-walk-template.md",
        """# Safety Walk Template

Checks: PPE, housekeeping, fall protection, hot work, ladder use.
""",
    )

    w(
        "rfi-template.md",
        """# RFI Template

RFI #:
Question:
Needed by:
Owner:
""",
    )

    w(
        "subcontractor-evaluation-template.md",
        """# Subcontractor Evaluation Template

Quality, schedule, safety, communication.
""",
    )


def make_json():
    wjson(
        "org-structure.json",
        {
            "generated": str(date.today()),
            "leaders": {
                "President": "Erin Caldwell",
                "VP Operations": "Mateo Reyes",
                "Precon": "Kira Olsen",
                "Controller": "Nolan Price",
                "Safety": "Leah Kim",
            },
            "offices": ["Phoenix", "Albuquerque"],
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-10-01",
            "kpis": [
                {"name": "CO total (approved)", "definition": "sum approved COs only"},
                {"name": "CO total (forecast)", "definition": "approved + pending"},
                {"name": "Margin (PM)", "definition": "PM forecast fee %"},
                {"name": "Margin (WIP)", "definition": "WIP margin %"},
            ],
        },
    )

    wjson(
        "owner-change-request-form.json",
        {
            "schema_version": "0.1",
            "title": "Owner Change Request",
            "fields": [
                {"id": "project", "type": "string", "required": True},
                {"id": "description", "type": "string", "required": True},
                {"id": "cost_impact", "type": "number"},
                {"id": "schedule_impact_days", "type": "integer"},
            ],
        },
    )


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "company-capabilities-one-pager.pdf",
        "Ironwood Ridge Constructors — Capabilities",
        ["Commercial", "Tenant improvements", "Healthcare buildouts", "Service"],
    )

    pdf(
        BASE / "contract-summary-ownerletter.pdf",
        "Contract Summary — Desert Vista (Owner Letter)",
        [
            "Contract type: GMP",
            "Change orders tracked as approved vs pending",
            "Schedule includes owner target dates and internal baselines",
        ],
    )


def make_xlsx():
    import openpyxl

    # estimate template
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Estimate"
    ws.append(["Division","Description","Amount"])
    for r in [("01","General conditions",2100000),("22","Plumbing",1850000),("23","HVAC",2400000),("26","Electrical",2650000),("09","Finishes",1980000)]:
        ws.append(list(r))
    for i in range(2, ws.max_row+1): ws.cell(i,3).number_format="$#,##0"
    wb.save(BASE/"estimate-template.xlsx")

    # master schedule
    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Master"
    ws2.append(["Milestone","Date","Notes"])
    ws2.append(["Substantial Completion","2026-11-05","internal schedule"])
    wb2.save(BASE/"master-schedule-desert-vista.xlsx")

    # lookahead
    wb3=openpyxl.Workbook(); ws3=wb3.active; ws3.title="Lookahead"
    ws3.append(["Week","Focus","Owner"])
    ws3.append(["2026-W41","MEP rough-in closeout","Super"])
    wb3.save(BASE/"lookahead-schedule-desert-vista.xlsx")

    # RFI log
    wb4=openpyxl.Workbook(); ws4=wb4.active; ws4.title="RFIs"
    ws4.append(["RFI #","Date","Question","Status","Owner"])
    ws4.append(["RFI-44","2026-09-03","Electrical room clearance","Open","PM"])
    wb4.save(BASE/"rfi-log-desert-vista.xlsx")

    # Submittal log
    wb5=openpyxl.Workbook(); ws5=wb5.active; ws5.title="Submittals"
    ws5.append(["Submittal","Date","Status","Owner"])
    ws5.append(["RTU cut sheets","2026-08-28","Revise","MEP Sub"])
    wb5.save(BASE/"submittal-log-desert-vista.xlsx")

    # CO log
    wb6=openpyxl.Workbook(); ws6=wb6.active; ws6.title="COs"
    ws6.append(["CO #","Date","Description","Cost","Status"])
    for r in [("CO-07","2026-08-22","Owner layout change",340000,"Approved"),("CO-09","2026-09-18","Medical gas add",400000,"Approved"),("CO-12","2026-09-30","Parking lot scope",310000,"Pending")]:
        ws6.append(list(r))
    for i in range(2, ws6.max_row+1): ws6.cell(i,4).number_format="$#,##0"
    wb6.save(BASE/"change-order-log-desert-vista.xlsx")

    # Job costing
    wb7=openpyxl.Workbook(); ws7=wb7.active; ws7.title="Job Cost"
    ws7.append(["Category","Budget","Cost to date","Forecast","Variance"])
    for r in [("MEP",6900000,4200000,7400000,None),("Finishes",1980000,860000,2120000,None)]:
        ws7.append(list(r))
    for i in range(2, ws7.max_row+1):
        ws7.cell(i,5).value=f"=B{i}-D{i}"
        for c in [2,3,4,5]: ws7.cell(i,c).number_format="$#,##0"
    wb7.save(BASE/"job-costing-desert-vista.xlsx")

    # WIP report
    wb8=openpyxl.Workbook(); ws8=wb8.active; ws8.title="WIP"
    ws8.append(["Project","Contract","CO Approved","Cost to Date","Est Cost at Comp","Margin % (WIP)"])
    ws8.append(["Desert Vista",22400000,740000,13800000,21050000,None])
    ws8.cell(2,6).value="=(B2+C2-E2)/(B2+C2)"; ws8.cell(2,6).number_format="0.0%"
    for c in [2,3,4,5]: ws8.cell(2,c).number_format="$#,##0"
    wb8.save(BASE/"wip-report-2026-09.xlsx")

    # AR
    wb9=openpyxl.Workbook(); ws9=wb9.active; ws9.title="AR"
    ws9.append(["Client","Invoice","Invoice Date","Amount","Status"])
    ws9.append(["Desert Vista Owner","APP-09","2026-09-30",1650000,"Open"])
    ws9.cell(2,4).number_format="$#,##0"
    wb9.save(BASE/"accounts-receivable-2026-09.xlsx")

    # Revenue
    wb10=openpyxl.Workbook(); ws10=wb10.active; ws10.title="Quarterly"
    ws10.append(["Year","Quarter","Revenue"])
    for r in [(2024,"Q4",54000000),(2025,"Q1",51000000),(2025,"Q2",56500000),(2025,"Q3",59000000),(2025,"Q4",62000000)]:
        ws10.append(list(r))
    for i in range(2, ws10.max_row+1): ws10.cell(i,3).number_format="$#,##0"
    wb10.save(BASE/"quarterly-revenue-history-2024-2025.xlsx")

    wb11=openpyxl.Workbook(); ws11=wb11.active; ws11.title="2026"
    ws11.append(["Month","Revenue"])
    for r in [("2026-01",18000000),("2026-02",17100000),("2026-03",19200000)]:
        ws11.append(list(r))
    for i in range(2, ws11.max_row+1): ws11.cell(i,2).number_format="$#,##0"
    wb11.save(BASE/"projected-revenue-2026.xlsx")

    # Safety incidents
    wb12=openpyxl.Workbook(); ws12=wb12.active; ws12.title="2026"
    ws12.append(["Date","Project","Incident","Severity"])
    ws12.append(["2026-09-21","Desert Vista","Heat stress near-miss","Low"])
    wb12.save(BASE/"safety-incident-log-2026.xlsx")

    # Punchlist template
    wb13=openpyxl.Workbook(); ws13=wb13.active; ws13.title="Template"
    ws13.append(["Area","Item","Owner","Status"])
    ws13.append(["Exam Room 2","Baseboard touch-up","Sub","Open"])
    wb13.save(BASE/"qaqc-punchlist-template.xlsx")


def main():
    make_markdown(); make_json(); make_pdfs(); make_xlsx()


if __name__ == "__main__":
    main()
