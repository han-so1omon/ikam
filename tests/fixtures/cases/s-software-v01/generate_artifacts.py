from __future__ import annotations

import json
from datetime import date
from pathlib import Path

BASE = Path(__file__).resolve().parent


def w(path: str, content: str):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content.rstrip() + "\n", encoding="utf-8")


def wjson(path: str, obj):
    p = BASE / path
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def make_md():
    w(
        "mission-vision-values.md",
        """# PocketRelay — Mission, Vision, Values

Mission: route customer requests to the right place instantly.
Vision: the simplest intake-and-routing layer for SMBs.
Values: small surface area, reliability, clear definitions.
""",
    )

    w(
        "brand-guide.md",
        """# PocketRelay — Brand Guide (v0.1)

Palette:
- Indigo: #312E81
- Mint: #34D399
- Off-white: #F8FAFC

Voice: friendly, direct.
""",
    )

    w(
        "one-page-strategy-2026.md",
        """# One-Page Strategy — 2026

- Stabilize webhooks and retries
- Simplify pricing tiers
- Ship Smart Routing v2
""",
    )

    w(
        "product-overview.md",
        """# Product Overview

PocketRelay is a lightweight intake layer: forms + webhooks + routing rules.

Smart Routing: rule-based routing with basic scoring.
""",
    )

    w(
        "mini-roadmap-2026.md",
        """# Mini Roadmap — 2026

- March: pricing cleanup
- April: Smart Routing v2
- Ongoing: webhook reliability
""",
    )

    w(
        "prd-smart-routing-v1.md",
        """# PRD — Smart Routing (v1)

Date: 2025-10-02
Notes: basic rules and simple filters.
""",
    )

    w(
        "prd-smart-routing-v2.md",
        """# PRD — Smart Routing (v2)

Date: 2026-02-08
Notes: adds scoring + routing audit log.
""",
    )

    w(
        "release-notes-2025.md",
        """# Release Notes — 2025

- v1 launch
- Webhook retries
- Early Smart Routing
""",
    )

    w(
        "pricing-page-copy.md",
        """# Pricing Page Copy

Starter: $49/mo
Team: $99/mo
Annual discount: 10%

Note: legacy tiers existed; some customers on old $79 tier.
""",
    )

    w(
        "support-macros.md",
        """# Support Macros

- "Confirm webhook URL"
- "Retry guidance"
- "How to export routing rules"
""",
    )

    w(
        "runbook-oncall.md",
        """# On-call Runbook

- Check queue depth
- Check worker lag
- Toggle retry backoff
""",
    )

    w(
        "incident-2026-01-17.md",
        """# Incident — 2026-01-17 — Webhook Retry Storm

Summary: webhook retries caused burst traffic.
Notes: write-up is incomplete; action items not fully captured.
""",
    )

    w(
        "founder-notes-2026-02-02.md",
        """# Founder Notes — 2026-02-02

- Pricing needs cleanup; contract addendum still mentions $79 tier.
- Churn definition debate: cancellations vs net churn.
""",
    )

    w(
        "customer-call-notes-2026-02-14.md",
        """# Customer Call Notes — 2026-02-14

Customer asked: "Are we on $79 or $99?"
Support replied using pricing page tiers.
""",
    )

    w(
        "incident-template.md",
        """# Incident Template

Summary:
Impact:
Timeline:
Root cause:
Actions:
""",
    )


def make_json():
    wjson(
        "org.json",
        {
            "generated": str(date.today()),
            "team": [
                {"name": "Riley Novak", "role": "Founder/CEO"},
                {"name": "Chen Wu", "role": "Founding Engineer"},
                {"name": "Maya Torres", "role": "Support/CS"},
                {"name": "Dev Patel", "role": "Fractional Finance"},
            ],
        },
    )

    wjson(
        "top-customers.json",
        {
            "generated": str(date.today()),
            "customers": [
                {"account": "Pinegrove HVAC", "plan": "Team"},
                {"account": "Cobalt Dental", "plan": "Starter (legacy addendum)"},
            ],
        },
    )

    wjson(
        "uptime-kpis.json",
        {
            "month": "2026-01",
            "uptime_pct": 99.72,
            "notes": "incident on 2026-01-17 affected webhook delivery latency",
        },
    )

    wjson(
        "kpi-definitions.json",
        {
            "version": "2026-02-02",
            "kpis": [
                {"name": "Churn (support)", "definition": "cancellations in period"},
                {"name": "Churn (finance)", "definition": "net revenue churn incl downgrades"},
                {"name": "Active", "definition": "account has >=1 routed request in 30 days"},
            ],
        },
    )

    wjson(
        "customer-intake-form.json",
        {
            "schema_version": "0.1",
            "title": "Customer Intake",
            "fields": [
                {"id": "business_name", "type": "string", "required": True},
                {"id": "industry", "type": "string"},
                {"id": "webhook_url", "type": "string"},
            ],
        },
    )


def make_xlsx():
    import openpyxl

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="2025"
    ws.append(["Month","MRR","Notes"])
    for r in [("2025-06",4200,"launch"),("2025-09",6100,""),("2025-12",7800,"")]:
        ws.append(list(r))
    for i in range(2, ws.max_row+1):
        ws.cell(i,2).number_format="$#,##0"
    wb.save(BASE/"revenue-history-2025.xlsx")

    wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="2026"
    ws2.append(["Month","MRR","Cancellations (count)","Downgrades ($)","Churn % (support)","Churn % (finance)"])
    ws2.append(["2026-01",8200,1,400,None,None])
    ws2.cell(2,5).value="=C2/40"  # fake denominator "accounts" in support narrative
    ws2.cell(2,6).value="=(C2*99 + D2)/B2"  # mixes cancellations and downgrade dollars
    ws2.cell(2,5).number_format="0.0%"; ws2.cell(2,6).number_format="0.0%"
    ws2.cell(2,2).number_format="$#,##0"; ws2.cell(2,4).number_format="$#,##0"
    wb2.save(BASE/"churn-and-mrr-2026.xlsx")


def make_pdfs():
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

    styles = getSampleStyleSheet()

    def pdf(out_path: Path, title: str, bullets: list[str]):
        doc = SimpleDocTemplate(str(out_path), pagesize=LETTER, title=title)
        story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
        items = [ListItem(Paragraph(b, styles["BodyText"])) for b in bullets]
        story.append(ListFlowable(items, bulletType="bullet"))
        doc.build(story)

    pdf(
        BASE / "msa-sample.pdf",
        "Master Services Agreement (Sample) — PocketRelay",
        [
            "Subscription service for intake + routing.",
            "Confidentiality and data handling.",
            "Pricing per order form.",
        ],
    )

    pdf(
        BASE / "contract-addendum-legacy-pricing.pdf",
        "Contract Addendum — Legacy Pricing",
        [
            "Legacy Team tier: $79/mo (superseded by pricing page).",
            "Customer: Cobalt Dental.",
        ],
    )


def main():
    make_md(); make_json(); make_xlsx(); make_pdfs()


if __name__ == "__main__":
    main()
