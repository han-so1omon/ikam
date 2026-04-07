"""XLSX structural diff engine with cell-level comparison.

Implements comparison of Excel workbooks at the cell level:
- Sheet additions/removals
- Cell value changes
- Formula changes with dependency tracking
- Optional: format changes (font, color, borders)

Performance: O(N*M) where N=sheets, M=cells per sheet (sparse iteration).
"""

from typing import Any, Dict, List, Optional, Union, Set
import re

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    # Allow module to load without openpyxl for testing
    pass

from .types import DiffChange, DiffResult


def _extract_cell_references(formula: str) -> Set[str]:
    """
    Extract cell references from a formula string.
    
    Args:
        formula: Formula string (without leading '=')
        
    Returns:
        Set of cell references (e.g., {'A1', 'B2', 'Sheet1!C3'})
        
    Examples:
        >>> _extract_cell_references("SUM(A1:A10)")
        {'A1', 'A10'}
        >>> _extract_cell_references("B2+C3*Sheet1!D4")
        {'B2', 'C3', 'Sheet1!D4'}
    """
    # Pattern for cell references: [Sheet!]A1 or [Sheet!]A1:B2
    # Supports both absolute ($A$1) and relative (A1) references
    pattern = r'(?:[A-Za-z_][\w\.]*!)?[$]?[A-Z]+[$]?\d+'
    
    matches = re.findall(pattern, formula)
    
    # Clean up and deduplicate
    refs = set()
    for match in matches:
        # Remove $ signs for canonical form
        clean = match.replace('$', '')
        refs.add(clean)
    
    return refs


def _formula_change_type(old_formula: str, new_formula: str) -> str:
    """
    Classify the type of formula change.
    
    Returns:
        - "formula_text": Formula text changed
        - "formula_dependencies": Cell references changed
        - "formula_function": Function name changed (e.g., SUM -> AVERAGE)
    """
    old_refs = _extract_cell_references(old_formula)
    new_refs = _extract_cell_references(new_formula)
    
    # Extract function names (uppercase words followed by '(')
    old_functions = set(re.findall(r'\b([A-Z]+)\s*\(', old_formula))
    new_functions = set(re.findall(r'\b([A-Z]+)\s*\(', new_formula))
    
    if old_functions != new_functions:
        return "formula_function"
    elif old_refs != new_refs:
        return "formula_dependencies"
    else:
        return "formula_text"


def compute_xlsx_diff(
    old_bytes: bytes,
    new_bytes: bytes,
    include_formulas: bool = True,
    include_formats: bool = False,
) -> DiffResult:
    """Compute structural diff between two Excel workbooks.
    
    Args:
        old_bytes: Original XLSX file bytes
        new_bytes: Updated XLSX file bytes
        include_formulas: Compare formula expressions (default True)
        include_formats: Compare cell formats like font, color (default False, MVP skip)
        
    Returns:
        DiffResult with cell-level changes using sheet!cell notation
        
    Examples:
        >>> with open("old.xlsx", "rb") as f:
        ...     old_bytes = f.read()
        >>> with open("new.xlsx", "rb") as f:
        ...     new_bytes = f.read()
        >>> result = compute_xlsx_diff(old_bytes, new_bytes)
        >>> result.changes[0].path
        'Sheet1!A1'
    """
    from io import BytesIO
    
    # Load workbooks
    old_wb = load_workbook(BytesIO(old_bytes), data_only=False)
    new_wb = load_workbook(BytesIO(new_bytes), data_only=False)
    
    changes: List[DiffChange] = []
    
    # Compare sheet structure
    old_sheets = set(old_wb.sheetnames)
    new_sheets = set(new_wb.sheetnames)
    
    # Added sheets
    for sheet_name in new_sheets - old_sheets:
        changes.append(DiffChange(
            path=f"{sheet_name}!<sheet>",
            change_type="added",
            new_value=f"Sheet '{sheet_name}' added",
        ))
    
    # Removed sheets
    for sheet_name in old_sheets - new_sheets:
        changes.append(DiffChange(
            path=f"{sheet_name}!<sheet>",
            change_type="removed",
            old_value=f"Sheet '{sheet_name}' removed",
        ))
    
    # Compare common sheets
    for sheet_name in old_sheets & new_sheets:
        old_sheet = old_wb[sheet_name]
        new_sheet = new_wb[sheet_name]
        _diff_sheets(old_sheet, new_sheet, sheet_name, changes, include_formulas, include_formats)
    
    return DiffResult(
        changes=changes,
        change_count=len(changes),
        affected_elements=len(set(c.path for c in changes)),
    )


def _diff_sheets(
    old_sheet: "Worksheet",
    new_sheet: "Worksheet",
    sheet_name: str,
    changes: List[DiffChange],
    include_formulas: bool,
    include_formats: bool,
) -> None:
    """Compare two worksheets cell-by-cell.
    
    Args:
        old_sheet: Original worksheet
        new_sheet: Updated worksheet
        sheet_name: Name of the sheet (for path construction)
        changes: Accumulator for discovered changes
        include_formulas: Whether to compare formulas
        include_formats: Whether to compare cell formats
    """
    # Collect all cell coordinates from both sheets
    old_cells = {(cell.row, cell.column): cell for row in old_sheet.iter_rows() for cell in row if cell.value is not None or (include_formulas and cell.data_type == 'f')}
    new_cells = {(cell.row, cell.column): cell for row in new_sheet.iter_rows() for cell in row if cell.value is not None or (include_formulas and cell.data_type == 'f')}
    
    old_coords = set(old_cells.keys())
    new_coords = set(new_cells.keys())
    
    # Added cells
    for coord in new_coords - old_coords:
        cell = new_cells[coord]
        cell_ref = cell.coordinate
        path = f"{sheet_name}!{cell_ref}"
        
        value_desc = _describe_cell(cell, include_formulas)
        changes.append(DiffChange(
            path=path,
            change_type="added",
            new_value=value_desc,
        ))
    
    # Removed cells
    for coord in old_coords - new_coords:
        cell = old_cells[coord]
        cell_ref = cell.coordinate
        path = f"{sheet_name}!{cell_ref}"
        
        value_desc = _describe_cell(cell, include_formulas)
        changes.append(DiffChange(
            path=path,
            change_type="removed",
            old_value=value_desc,
        ))
    
    # Modified cells
    for coord in old_coords & new_coords:
        old_cell = old_cells[coord]
        new_cell = new_cells[coord]
        cell_ref = old_cell.coordinate
        path = f"{sheet_name}!{cell_ref}"
        
        old_is_formula = old_cell.data_type == 'f'
        new_is_formula = new_cell.data_type == 'f'
        
        # Case 1: Both are formulas - compare formula text
        if old_is_formula and new_is_formula and include_formulas:
            old_formula = str(old_cell.value) if old_cell.value else ""
            new_formula = str(new_cell.value) if new_cell.value else ""
            
            if old_formula != new_formula:
                # Determine change type
                change_subtype = _formula_change_type(old_formula, new_formula)
                
                # Extract dependencies for metadata
                old_refs = _extract_cell_references(old_formula)
                new_refs = _extract_cell_references(new_formula)
                
                change = DiffChange(
                    path=path,
                    change_type="modified",
                    old_value=f"={old_formula}",
                    new_value=f"={new_formula}",
                )
                # Store metadata about formula change
                change.__dict__['formula_change_type'] = change_subtype
                change.__dict__['old_dependencies'] = sorted(old_refs)
                change.__dict__['new_dependencies'] = sorted(new_refs)
                changes.append(change)
        
        # Case 2: Formula converted to value or vice versa
        elif old_is_formula != new_is_formula:
            changes.append(DiffChange(
                path=path,
                change_type="modified",
                old_value=_describe_cell(old_cell, include_formulas),
                new_value=_describe_cell(new_cell, include_formulas),
            ))
        
        # Case 3: Both are values - simple value comparison
        elif old_cell.value != new_cell.value:
            changes.append(DiffChange(
                path=path,
                change_type="modified",
                old_value=_describe_cell(old_cell, include_formulas),
                new_value=_describe_cell(new_cell, include_formulas),
            ))
        
        # TODO: Compare formats if include_formats=True (font, fill, border, number_format)
        # For MVP, skip format comparison


def _describe_cell(cell: Any, include_formulas: bool) -> Union[str, int, float, None]:
    """Describe a cell's value for diff output.
    
    Args:
        cell: openpyxl Cell object
        include_formulas: Whether to include formula expressions
        
    Returns:
        Cell value or formula string
    """
    if cell.data_type == 'f' and include_formulas:
        # Formula cell - return formula expression
        return f"={cell.value}"
    else:
        # Regular value
        return cell.value
