"""Excel rendering utilities with deterministic mode support.

Phase 4 deterministic export infrastructure:
- Frozen timestamps
- Stable IDs (deterministic sheet naming ordering)
- Float precision rounding

The renderer inspects IKAM_* environment variables:
- IKAM_DETERMINISTIC_RENDER (bool)
- IKAM_FROZEN_TIMESTAMP (ISO8601 string)
- IKAM_STABLE_IDS (bool)
- IKAM_FLOAT_PRECISION (int)

Usage:
    from ikam.renderers.excel import render_excel
    content_bytes = render_excel({"sheets": [{"name": "Financials", "cells": {"A1": 1.123456789}}]})
"""
from __future__ import annotations

import os
from io import BytesIO
from datetime import datetime
import uuid
from typing import Any, Dict

from openpyxl import Workbook


def _env_bool(key: str, default: str = "false") -> bool:
    return os.getenv(key, default).lower() in {"1", "true", "yes"}


def _env_int(key: str, default: str) -> int:
    try:
        return int(os.getenv(key, default))
    except ValueError:
        return int(default)


def _parse_timestamp(raw: str | None) -> datetime | None:
    if not raw:
        return None
    # Support trailing Z
    if raw.endswith("Z"):
        raw = raw.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def render_excel(model: Dict[str, Any]) -> bytes:
    """Render a simple Excel workbook from a model structure.

    Model format:
        {
          "sheets": [
             {"name": "Sheet Name", "cells": {"A1": value, "B2": value, ...}},
             ...
          ]
        }

    Deterministic behavior controlled by IKAM_* env variables.
    """
    deterministic = _env_bool("IKAM_DETERMINISTIC_RENDER")
    frozen_ts_raw = os.getenv("IKAM_FROZEN_TIMESTAMP")
    frozen_dt = _parse_timestamp(frozen_ts_raw) if deterministic else None
    stable_ids = _env_bool("IKAM_STABLE_IDS") if deterministic else False
    float_precision = _env_int("IKAM_FLOAT_PRECISION", "15") if deterministic else 15

    wb = Workbook()

    # Remove default sheet if we will re-create deterministically
    default_sheet = wb.active

    sheets = model.get("sheets", [])

    # Deterministic ordering: sort by provided name if stable_ids enabled
    if stable_ids:
        sheets = sorted(sheets, key=lambda s: s.get("name", ""))

    # If first sheet mapping should reuse default sheet, rename it deterministically.
    if sheets:
        first = sheets[0]
        sheet_name = first.get("name") or "Sheet1"
        default_sheet.title = sheet_name
        _populate_sheet(default_sheet, first.get("cells", {}), float_precision, deterministic)
        # add remaining
        for sheet in sheets[1:]:
            ws = wb.create_sheet(title=sheet.get("name") or f"Sheet{len(wb.worksheets)}")
            _populate_sheet(ws, sheet.get("cells", {}), float_precision, deterministic)
    else:
        default_sheet.title = "Sheet1" if stable_ids else default_sheet.title

    if deterministic:
        _apply_deterministic_properties(wb, frozen_dt)
    else:
        # Introduce intentional non-determinism so snapshots detect variance.
        token = uuid.uuid4().hex
        default_sheet["ZZ1"].value = token

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _populate_sheet(ws, cells: Dict[str, Any], float_precision: int, deterministic: bool) -> None:
    for addr, value in cells.items():
        # Handle cell objects: {"value": X, "type": "formula"} or {"value": X}
        if isinstance(value, dict):
            if "value" in value:
                cell_value = value["value"]
                # Set formula or raw value
                ws[addr].value = cell_value
            else:
                # Unknown dict format, skip
                pass
        elif isinstance(value, float) and deterministic:
            rounded = round(value, float_precision)
            ws[addr].value = rounded
        else:
            ws[addr].value = value


def _apply_deterministic_properties(wb: Workbook, frozen_dt: datetime | None) -> None:
    props = wb.properties
    # openpyxl DocumentProperties may not expose direct assignment for all fields
    if frozen_dt:
        # Format to RFC3339-style string without microseconds
        ts = frozen_dt.replace(microsecond=0)
        props.created = ts
        props.modified = ts
    # Clear volatile attributes (if present)
    _safe_clear_attr(props, "lastPrinted")
    _safe_clear_attr(props, "category")


def _safe_clear_attr(props, attr: str) -> None:
    if hasattr(props, attr):
        try:
            setattr(props, attr, None)
        except Exception:
            pass

__all__ = ["render_excel"]
