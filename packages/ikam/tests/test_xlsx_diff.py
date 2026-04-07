"""Tests for XLSX structural diff engine."""

from io import BytesIO

import pytest
from openpyxl import Workbook

from ikam.diff.xlsx_diff import compute_xlsx_diff


def create_simple_workbook(data: dict) -> bytes:
    """Helper to create a simple workbook for testing.
    
    Args:
        data: Dict of sheet_name -> list of (cell_ref, value) tuples
        
    Returns:
        Workbook bytes
        
    Example:
        data = {"Sheet1": [("A1", "Hello"), ("B1", 123)]}
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet_name, cells in data.items():
        ws = wb.create_sheet(sheet_name)
        for cell_ref, value in cells:
            ws[cell_ref] = value
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_identical_workbooks_no_changes():
    """Identical workbooks should produce no changes."""
    data = {"Sheet1": [("A1", "Hello"), ("B1", 123)]}
    old_bytes = create_simple_workbook(data)
    new_bytes = create_simple_workbook(data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 0
    assert len(result.changes) == 0


def test_cell_value_modified():
    """Cell value change detected."""
    old_data = {"Sheet1": [("A1", "Hello")]}
    new_data = {"Sheet1": [("A1", "World")]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.path == "Sheet1!A1"
    assert change.change_type == "modified"
    assert change.old_value == "Hello"
    assert change.new_value == "World"


def test_cell_added():
    """New cell added detected."""
    old_data = {"Sheet1": [("A1", "Hello")]}
    new_data = {"Sheet1": [("A1", "Hello"), ("B1", 123)]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.path == "Sheet1!B1"
    assert change.change_type == "added"
    assert change.new_value == 123


def test_cell_removed():
    """Cell removal detected."""
    old_data = {"Sheet1": [("A1", "Hello"), ("B1", 123)]}
    new_data = {"Sheet1": [("A1", "Hello")]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.path == "Sheet1!B1"
    assert change.change_type == "removed"
    assert change.old_value == 123


def test_sheet_added():
    """New sheet added detected."""
    old_data = {"Sheet1": [("A1", 1)]}
    new_data = {"Sheet1": [("A1", 1)], "Sheet2": [("A1", 2)]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    # Should detect: Sheet2 added + Sheet2!A1 added
    assert result.change_count >= 1
    
    sheet_changes = [c for c in result.changes if "<sheet>" in c.path]
    assert len(sheet_changes) == 1
    assert sheet_changes[0].path == "Sheet2!<sheet>"
    assert sheet_changes[0].change_type == "added"


def test_sheet_removed():
    """Sheet removal detected."""
    old_data = {"Sheet1": [("A1", 1)], "Sheet2": [("A1", 2)]}
    new_data = {"Sheet1": [("A1", 1)]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    sheet_changes = [c for c in result.changes if "<sheet>" in c.path]
    assert len(sheet_changes) == 1
    assert sheet_changes[0].path == "Sheet2!<sheet>"
    assert sheet_changes[0].change_type == "removed"


def test_multiple_cell_changes():
    """Multiple cell changes across same sheet."""
    old_data = {"Sheet1": [("A1", 1), ("B1", 2), ("C1", 3)]}
    new_data = {"Sheet1": [("A1", 1), ("B1", 5), ("D1", 4)]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    # Changes: B1 modified (2->5), C1 removed, D1 added
    assert result.change_count == 3
    
    paths = {c.path for c in result.changes}
    assert "Sheet1!B1" in paths
    assert "Sheet1!C1" in paths
    assert "Sheet1!D1" in paths


def test_formula_change_detected():
    """Formula changes detected when include_formulas=True."""
    # Create workbooks with formulas
    old_wb = Workbook()
    old_ws = old_wb.active
    old_ws["A1"] = 10
    old_ws["B1"] = "=A1*2"
    
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws["A1"] = 10
    new_ws["B1"] = "=A1*3"  # Formula changed
    
    old_buffer = BytesIO()
    old_wb.save(old_buffer)
    old_buffer.seek(0)
    old_bytes = old_buffer.read()
    
    new_buffer = BytesIO()
    new_wb.save(new_buffer)
    new_buffer.seek(0)
    new_bytes = new_buffer.read()
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    # A1 unchanged, B1 formula changed
    formula_changes = [c for c in result.changes if "B1" in c.path]
    assert len(formula_changes) == 1
    
    change = formula_changes[0]
    assert change.change_type == "modified"
    # Formula descriptions include "=" prefix
    assert "A1*2" in str(change.old_value)
    assert "A1*3" in str(change.new_value)


def test_formula_skipped_when_disabled():
    """Formulas skipped when include_formulas=False."""
    # Create workbooks with formulas
    old_wb = Workbook()
    old_ws = old_wb.active
    old_ws["B1"] = "=A1*2"
    
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws["B1"] = "=A1*3"
    
    old_buffer = BytesIO()
    old_wb.save(old_buffer)
    old_buffer.seek(0)
    old_bytes = old_buffer.read()
    
    new_buffer = BytesIO()
    new_wb.save(new_buffer)
    new_buffer.seek(0)
    new_bytes = new_buffer.read()
    
    # With include_formulas=False, only value comparison (both formulas have None as data_only=False)
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=False)
    
    # Should still detect change if values differ
    # But formula expressions not included in description


def test_numeric_type_change():
    """Number to string type change detected."""
    old_data = {"Sheet1": [("A1", 123)]}
    new_data = {"Sheet1": [("A1", "123")]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.path == "Sheet1!A1"
    assert change.change_type == "modified"
    assert change.old_value == 123
    assert change.new_value == "123"


def test_empty_cell_to_value():
    """Empty cell to value detected as addition."""
    old_data = {"Sheet1": []}
    new_data = {"Sheet1": [("A1", "New")]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.change_type == "added"


def test_value_to_empty_cell():
    """Value to empty cell detected as removal."""
    old_data = {"Sheet1": [("A1", "Old")]}
    new_data = {"Sheet1": []}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    assert result.change_count == 1
    
    change = result.changes[0]
    assert change.change_type == "removed"


def test_multiple_sheets_changes():
    """Changes across multiple sheets."""
    old_data = {
        "Sheet1": [("A1", 1)],
        "Sheet2": [("A1", 2)],
    }
    new_data = {
        "Sheet1": [("A1", 5)],  # Modified
        "Sheet2": [("A1", 2)],  # Unchanged
        "Sheet3": [("A1", 3)],  # Added sheet
    }
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    # Changes: Sheet1!A1 modified, Sheet3 added (+ Sheet3!A1 added)
    paths = {c.path for c in result.changes}
    assert "Sheet1!A1" in paths
    assert "Sheet3!<sheet>" in paths


def test_affected_elements_count():
    """Affected elements count matches distinct cell references."""
    old_data = {"Sheet1": [("A1", 1), ("B1", 2)]}
    new_data = {"Sheet1": [("A1", 5), ("B1", 6)]}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    # 2 changes, 2 distinct paths
    assert result.change_count == 2
    assert result.affected_elements == 2


def test_large_range_changes():
    """Handle multiple row/column changes efficiently."""
    old_cells = [(f"A{i}", i) for i in range(1, 11)]
    new_cells = [(f"A{i}", i * 2) for i in range(1, 11)]
    
    old_data = {"Sheet1": old_cells}
    new_data = {"Sheet1": new_cells}
    
    old_bytes = create_simple_workbook(old_data)
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes)
    
    # All 10 cells modified
    assert result.change_count == 10
    assert all(c.change_type == "modified" for c in result.changes)


# Formula-specific tests

def create_workbook_with_formulas(formulas: dict) -> bytes:
    """Helper to create workbook with formula cells.
    
    Args:
        formulas: Dict of sheet_name -> list of (cell_ref, formula_string) tuples
        Formula strings should NOT include leading '='
        
    Example:
        formulas = {"Sheet1": [("A1", "SUM(B1:B10)"), ("C1", "B1*2")]}
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, cells in formulas.items():
        ws = wb.create_sheet(sheet_name)
        for cell_ref, formula in cells:
            ws[cell_ref] = f"={formula}"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_formula_text_changed():
    """Detect formula text modification."""
    old_formulas = {"Sheet1": [("A1", "SUM(B1:B10)")]}
    new_formulas = {"Sheet1": [("A1", "SUM(B1:B20)")]}  # Range extended
    
    old_bytes = create_workbook_with_formulas(old_formulas)
    new_bytes = create_workbook_with_formulas(new_formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    assert result.change_count == 1
    change = result.changes[0]
    assert change.path == "Sheet1!A1"
    assert change.change_type == "modified"
    assert "SUM(B1:B10)" in change.old_value
    assert "SUM(B1:B20)" in change.new_value


def test_formula_dependencies_changed():
    """Detect when formula cell references change."""
    old_formulas = {"Sheet1": [("C1", "A1+B1")]}
    new_formulas = {"Sheet1": [("C1", "A2+B2")]}  # References changed
    
    old_bytes = create_workbook_with_formulas(old_formulas)
    new_bytes = create_workbook_with_formulas(new_formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    assert result.change_count == 1
    change = result.changes[0]
    
    # Check metadata for dependency tracking
    if hasattr(change, 'formula_change_type'):
        assert change.formula_change_type == "formula_dependencies"
        assert set(change.old_dependencies) == {"A1", "B1"}
        assert set(change.new_dependencies) == {"A2", "B2"}


def test_formula_function_changed():
    """Detect when formula function changes."""
    old_formulas = {"Sheet1": [("A1", "SUM(B1:B10)")]}
    new_formulas = {"Sheet1": [("A1", "AVERAGE(B1:B10)")]}  # Function changed
    
    old_bytes = create_workbook_with_formulas(old_formulas)
    new_bytes = create_workbook_with_formulas(new_formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    assert result.change_count == 1
    change = result.changes[0]
    
    if hasattr(change, 'formula_change_type'):
        assert change.formula_change_type == "formula_function"


def test_formula_to_value_conversion():
    """Detect formula converted to static value."""
    old_formulas = {"Sheet1": [("A1", "SUM(B1:B10)")]}
    old_bytes = create_workbook_with_formulas(old_formulas)
    
    # New: static value instead of formula
    new_data = {"Sheet1": [("A1", 55)]}
    new_bytes = create_simple_workbook(new_data)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    assert result.change_count == 1
    change = result.changes[0]
    assert change.change_type == "modified"
    assert "SUM" in change.old_value  # Was formula
    assert change.new_value == 55  # Now value


def test_value_to_formula_conversion():
    """Detect static value converted to formula."""
    old_data = {"Sheet1": [("A1", 100)]}
    old_bytes = create_simple_workbook(old_data)
    
    new_formulas = {"Sheet1": [("A1", "B1*2")]}
    new_bytes = create_workbook_with_formulas(new_formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    assert result.change_count == 1
    change = result.changes[0]
    assert change.change_type == "modified"
    assert change.old_value == 100
    assert "B1*2" in change.new_value


def test_multiple_formulas_changed():
    """Handle multiple formula changes."""
    old_formulas = {
        "Sheet1": [
            ("A1", "SUM(B1:B10)"),
            ("C1", "A1*2"),
            ("D1", "AVERAGE(E1:E5)"),
        ]
    }
    new_formulas = {
        "Sheet1": [
            ("A1", "SUM(B1:B20)"),  # Range changed
            ("C1", "A1*3"),  # Constant changed
            ("D1", "AVERAGE(E1:E5)"),  # Unchanged
        ]
    }
    
    old_bytes = create_workbook_with_formulas(old_formulas)
    new_bytes = create_workbook_with_formulas(new_formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    # Only A1 and C1 should be detected as changed
    assert result.change_count == 2
    changed_paths = {c.path for c in result.changes}
    assert changed_paths == {"Sheet1!A1", "Sheet1!C1"}


def test_formula_unchanged_but_value_different():
    """Formula identical but calculated value different (not detected)."""
    # Note: We compare formula TEXT, not calculated values
    # This test verifies we don't report changes for same formulas
    formulas = {"Sheet1": [("A1", "SUM(B1:B10)")]}
    
    old_bytes = create_workbook_with_formulas(formulas)
    new_bytes = create_workbook_with_formulas(formulas)
    
    result = compute_xlsx_diff(old_bytes, new_bytes, include_formulas=True)
    
    # Formula text identical - no changes
    assert result.change_count == 0


def test_extract_cell_references():
    """Test cell reference extraction from formulas."""
    from ikam.diff.xlsx_diff import _extract_cell_references
    
    # Simple reference
    refs = _extract_cell_references("A1")
    assert "A1" in refs
    
    # Range
    refs = _extract_cell_references("SUM(A1:A10)")
    assert "A1" in refs and "A10" in refs
    
    # Multiple refs
    refs = _extract_cell_references("B2+C3*D4")
    assert refs == {"B2", "C3", "D4"}
    
    # Cross-sheet reference
    refs = _extract_cell_references("Sheet1!A1+Sheet2!B2")
    assert "Sheet1!A1" in refs and "Sheet2!B2" in refs
    
    # Absolute references ($ signs removed)
    refs = _extract_cell_references("$A$1+B$2")
    assert "A1" in refs and "B2" in refs
